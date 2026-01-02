import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
import os
import winreg

# 날짜 범위 설정
start_date = datetime(2025, 11, 17)
end_date = datetime(2025, 11, 28)

# 엑셀 워크북 생성
wb = openpyxl.Workbook()

# 기본 시트 삭제
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# ========== 스타일 정의 ==========
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
date_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
lunch_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# 카테고리별 색상
category_colors = {
    "시스템/프로그램": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "업무 프로세스": PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"),
    "고객사/거래처": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "문서/매뉴얼": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
    "기타": PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
}

# 사이드바 스타일
sidebar_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
sidebar_border = Border(
    right=Side(style='medium', color='CCCCCC')
)

# ========== 시트 그룹 정의 (논리적 그룹핑) ==========
sheet_groups = [
    {
        "name": "일정 및 진행 관리",
        "color": "DEEBF7",  # 연한 파랑
        "sheets": ["워크플로우", "인수인계 스케줄", "할 일 목록"]
    },
    {
        "name": "업무 가이드",
        "color": "E2EFDA",  # 연한 초록
        "sheets": ["고객처별가이드", "SAP사용법", "시스템도구"]
    },
    {
        "name": "인적/조직 정보",
        "color": "FFF2CC",  # 연한 노랑
        "sheets": ["담당자 정보", "고객사거래처"]
    },
    {
        "name": "참고 자료",
        "color": "F4B084",  # 연한 주황
        "sheets": ["참고자료", "용어집"]
    },
    {
        "name": "기술문서",
        "color": "E7E6E6",  # 회색
        "sheets": ["구현가이드"]
    }
]

# ========== 시트 목록 정의 ==========
sheets_info = [
    {"name": "워크플로우", "icon": ""},
    {"name": "구현가이드", "icon": ""},
    {"name": "인수인계 스케줄", "icon": ""},
    {"name": "할 일 목록", "icon": ""},
    {"name": "참고자료", "icon": ""},
    {"name": "담당자 정보", "icon": ""},
    {"name": "고객사거래처", "icon": ""},
    {"name": "SAP사용법", "icon": ""},
    {"name": "시스템도구", "icon": ""},
    {"name": "용어집", "icon": ""},
    {"name": "고객처별가이드", "icon": ""}
]

# 사이드바 생성 함수 (개선된 버전 - 그룹핑 및 시각적 개선)
def create_sidebar(ws, current_sheet_name):
    """각 시트에 사이드바 네비게이션 생성 (그룹핑 및 시각적 개선)"""
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 5
    
    # 사이드바 배경
    for row in range(1, 100):
        cell = ws.cell(row=row, column=1)
        cell.fill = sidebar_fill
        cell.border = sidebar_border
    
    # 네비게이션 제목 (강조)
    nav_title = ws.cell(row=1, column=1, value="네비게이션")
    nav_title.font = Font(bold=True, size=13, color="FFFFFF")
    nav_title.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    nav_title.border = sidebar_border
    nav_title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    row = 2
    
    # 각 그룹별로 시트 나열
    for group in sheet_groups:
        # 그룹 헤더
        group_header = ws.cell(row=row, column=1, value=group["name"])
        group_header.font = Font(bold=True, size=11, color="000000")
        group_header.fill = PatternFill(start_color=group["color"], end_color=group["color"], fill_type="solid")
        group_header.border = sidebar_border
        group_header.alignment = Alignment(horizontal='left', vertical='center', indent=0)
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 그룹 내 시트들
        for sheet_name in group["sheets"]:
            # 시트 정보 찾기
            sheet_info = None
            for s in sheets_info:
                if s["name"] == sheet_name:
                    sheet_info = s
                    break
            
            if sheet_info:
                display_name = sheet_name
                
                # 자주 사용하는 시트 표시
                frequent_sheets = ["인수인계 스케줄", "할 일 목록"]
                star = "" if sheet_name in frequent_sheets else ""
                
                cell = ws.cell(row=row, column=1, value=f"  {display_name}{star}")
                cell.font = Font(size=10)
                cell.fill = sidebar_fill
                cell.border = sidebar_border
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
                
                # 현재 시트 강조
                if sheet_name == current_sheet_name:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    cell.font = Font(size=10, bold=True, color="000000")
                    cell.border = Border(
                        left=Side(style='thick', color='366092'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                else:
                    # 하이퍼링크 추가
                    cell.hyperlink = f"#'{sheet_name}'!A1"
                    cell.font = Font(size=10, color="0000FF", underline="single")
                
                row += 1
        
        # 그룹 간 구분선
        if group != sheet_groups[-1]:  # 마지막 그룹이 아니면
            ws.cell(row=row, column=1, value="").fill = sidebar_fill
            ws.cell(row=row, column=1).border = Border(
                bottom=Side(style='thin', color='CCCCCC')
            )
            row += 1
    
# ========== 1. 인수인계 스케줄 시트 ==========
schedule_ws = wb.create_sheet("인수인계 스케줄")
create_sidebar(schedule_ws, "인수인계 스케줄")

# 헤더 작성
headers = [
    "날짜", "요일", "시간", "주제/내용", "담당자", 
    "소요시간", "상세내용", "출근자", "재택근무자", "비고", "완료여부"
]

header_start_col = 3
for col_idx, header in enumerate(headers, start=header_start_col):
    cell = schedule_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# 날짜별로 행 생성
row = 2
current_date = start_date
weekdays = ['월', '화', '수', '목', '금', '토', '일']

while current_date <= end_date:
    weekday = weekdays[current_date.weekday()]
    
    # 시간대별 행 생성 (9시~18시, 1시간 단위)
    time_slots = []
    for hour in range(9, 18):
        time_str = f"{hour:02d}:00 - {hour+1:02d}:00"
        time_slots.append(time_str)
    
    lunch_time_idx = 3  # 12:00-13:00
    
    for time_idx, time_slot in enumerate(time_slots):
        current_row = row + time_idx
        
        # 날짜와 요일 (첫 번째 시간대에만)
        if time_idx == 0:
            date_cell = schedule_ws.cell(row=current_row, column=header_start_col, 
                                         value=current_date.strftime("%Y-%m-%d"))
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            date_cell.border = border_style
            date_cell.font = Font(bold=True, size=10)
            date_cell.fill = date_fill
            
            weekday_cell = schedule_ws.cell(row=current_row, column=header_start_col + 1, value=weekday)
            weekday_cell.alignment = Alignment(horizontal='center', vertical='center')
            weekday_cell.border = border_style
            weekday_cell.fill = date_fill
        else:
            schedule_ws.cell(row=current_row, column=header_start_col, value="").border = border_style
            schedule_ws.cell(row=current_row, column=header_start_col + 1, value="").border = border_style
        
        is_lunch = (time_idx == lunch_time_idx)
        
        # 시간
        time_cell = schedule_ws.cell(row=current_row, column=header_start_col + 2, value=time_slot)
        time_cell.alignment = Alignment(horizontal='center', vertical='center')
        time_cell.border = border_style
        if is_lunch:
            time_cell.fill = lunch_fill
            time_cell.value = "12:00 - 13:00 (점심시간)"
        
        # 나머지 셀들
        for col_offset in range(3, len(headers)):
            cell = schedule_ws.cell(row=current_row, column=header_start_col + col_offset, value="")
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border_style
            if is_lunch:
                cell.fill = lunch_fill
    
    # 날짜와 요일 셀 병합
    if len(time_slots) > 1:
        schedule_ws.merge_cells(f'{get_column_letter(header_start_col)}{row}:{get_column_letter(header_start_col)}{row + len(time_slots) - 1}')
        schedule_ws.merge_cells(f'{get_column_letter(header_start_col + 1)}{row}:{get_column_letter(header_start_col + 1)}{row + len(time_slots) - 1}')
    
    row += len(time_slots)
    current_date += timedelta(days=1)

# 열 너비 조정
column_widths = {'C': 12, 'D': 8, 'E': 18, 'F': 35, 'G': 20, 'H': 12, 'I': 40, 'J': 18, 'K': 18, 'L': 25, 'M': 12}
for col, width in column_widths.items():
    schedule_ws.column_dimensions[col].width = width

# 행 높이 조정
for row in range(2, schedule_ws.max_row + 1):
    schedule_ws.row_dimensions[row].height = 30
schedule_ws.row_dimensions[1].height = 35

# 완료여부 드롭다운
status_validation = DataValidation(type="list", formula1='"미완료,진행중,완료,보류"', allow_blank=True)
schedule_ws.add_data_validation(status_validation)
status_validation.add(f'M2:M{schedule_ws.max_row}')

# 필터 추가
schedule_ws.auto_filter.ref = f'C1:{get_column_letter(schedule_ws.max_column)}{schedule_ws.max_row}'

# 조건부 서식: 오늘 날짜 하이라이트 (날짜 컬럼 C)
from openpyxl.formatting.rule import FormulaRule
today_rule = FormulaRule(formula=['$C2=TODAY()'], stopIfTrue=True, 
                        fill=PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"))
schedule_ws.conditional_formatting.add(f'C2:M{schedule_ws.max_row}', today_rule)

# 조건부 서식: 완료여부별 색상 (완료여부 컬럼 M)
complete_rule = CellIsRule(operator='equal', formula=['"완료"'], 
                          fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
schedule_ws.conditional_formatting.add(f'M2:M{schedule_ws.max_row}', complete_rule)

progress_rule = CellIsRule(operator='equal', formula=['"진행중"'], 
                          fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
schedule_ws.conditional_formatting.add(f'M2:M{schedule_ws.max_row}', progress_rule)

pending_rule = CellIsRule(operator='equal', formula=['"미완료"'], 
                         fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
schedule_ws.conditional_formatting.add(f'M2:M{schedule_ws.max_row}', pending_rule)

# ========== 3. 할 일 목록 시트 ==========
todo_ws = wb.create_sheet("할 일 목록")
create_sidebar(todo_ws, "할 일 목록")

# 헤더 (체크리스트 기능 통합: 단계 컬럼 추가)
todo_headers = ["카테고리", "할 일 항목", "단계", "담당자", "우선순위", "예상소요시간", "상세내용", "완료여부", "비고"]
for col_idx, header in enumerate(todo_headers, start=3):
    cell = todo_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# 할 일 데이터
tasks = [
    {"category": "시스템/프로그램", "task": "STARNET 등 설치파일 설치", "person": "정인님", 
     "priority": "높음", "duration": "1시간", "detail": "STARNET 등 필요한 프로그램 설치 및 설정 확인", 
     "status": "", "note": "설치 파일 위치 및 설치 순서 매뉴얼 준비"},
    {"category": "시스템/프로그램", "task": "서명 만들기", "person": "정인님", 
     "priority": "중", "duration": "30분", "detail": "이메일 서명 파일 생성 및 설정", 
     "status": "", "note": "서명 템플릿 파일 전달 필요"},
    {"category": "업무 프로세스", "task": "정인님 인수인계 일정 및 내용 설명", "person": "담당자", 
     "priority": "높음", "duration": "1시간", "detail": "전체 인수인계 일정 및 각 항목별 내용 설명", 
     "status": "", "note": "첫날 오전에 진행"},
    {"category": "고객사/거래처", "task": "침퓨즈(APRO, DCC) 인수인계", "person": "영화 과장님", 
     "priority": "높음", "duration": "40분", "detail": "침퓨즈 관련 APRO, DCC 업무 프로세스 및 주요 사항 인수인계", 
     "status": "", "note": "출근자: 영화, 송은, 현정"},
    {"category": "고객사/거래처", "task": "악세스(ACCSS-노이즈필터) 인수인계", "person": "승은 차장님, 현정 과장님", 
     "priority": "높음", "duration": "20분", "detail": "ACCSS 노이즈필터 관련 업무 프로세스 인수인계", 
     "status": "", "note": "정인님 OJT"},
    {"category": "고객사/거래처", "task": "커넥터 EMS 3사 전체적으로 인수인계 설명", "person": "미경 선배님", 
     "priority": "높음", "duration": "1시간", "detail": "커넥터 EMS 3개사 전체적인 업무 프로세스 및 주요 사항 설명", 
     "status": "", "note": "출근자: 미경, 현정"},
    {"category": "고객사/거래처", "task": "한국성전 - FCST 인계", "person": "미경 선배님", 
     "priority": "중", "duration": "30분", "detail": "한국성전 FCST 관련 업무 인수인계", 
     "status": "", "note": "출근자: 미경, 현정"},
    {"category": "고객사/거래처", "task": "샘플 인계 - 유상 및 발주 부분", "person": "담당자", 
     "priority": "중", "duration": "1시간", "detail": "샘플 관련 유상 및 발주 업무 프로세스 인수인계", 
     "status": "", "note": "출근자: 송은"},
    {"category": "고객사/거래처", "task": "샘플 - 발주 이레귤러 바뀐 부분 인계", "person": "미경 선배님", 
     "priority": "중", "duration": "30분", "detail": "샘플 발주 중 이레귤러하게 변경된 부분에 대한 인수인계", 
     "status": "", "note": "출근자: 송은, 현정, 미경"},
    {"category": "문서/매뉴얼", "task": "매뉴얼 파일 정리 및 전달", "person": "담당자", 
     "priority": "높음", "duration": "2시간", "detail": "정리한 매뉴얼 파일을 정인님에게 전달하여 업무 수행하게 하기", 
     "status": "", "note": "각 업무별 매뉴얼 파일 체계적으로 정리 필요"},
    {"category": "문서/매뉴얼", "task": "할 일 시트 링크 생성", "person": "담당자", 
     "priority": "낮음", "duration": "30분", "detail": "옆에 할 일에 대한 시트로 넘어가는 링크 만들기", 
     "status": "", "note": "엑셀 하이퍼링크 기능 활용"},
    {"category": "기타", "task": "출근자 일정 확인 및 조율", "person": "담당자", 
     "priority": "중", "duration": "1시간", "detail": "각 인수인계 항목별 필요한 출근자 일정 확인 및 조율", 
     "status": "", "note": "재택근무자와 출근자 일정 매칭 필요"},
    {"category": "기타", "task": "인수인계 완료 체크리스트 작성", "person": "담당자", 
     "priority": "중", "duration": "1시간", "detail": "각 항목별 인수인계 완료 여부 확인용 체크리스트 작성", 
     "status": "", "note": "실제 업무 수행 가능 여부 확인"},
    {"category": "기타", "task": "Q&A 시간 확보", "person": "전체", 
     "priority": "중", "duration": "매일 30분", "detail": "인수인계 후 궁금한 사항 질문 및 답변 시간", 
     "status": "", "note": "매일 하루 종료 전 정리 시간 활용"}
]

# 할 일 목록 작성
todo_row = 2
for task in tasks:
    col = 3
    # 카테고리
    cat_cell = todo_ws.cell(row=todo_row, column=col, value=task["category"])
    cat_cell.fill = category_colors.get(task["category"], category_colors["기타"])
    cat_cell.alignment = Alignment(horizontal='center', vertical='center')
    cat_cell.border = border_style
    cat_cell.font = Font(size=10)
    col += 1
    
    # 할 일 항목
    task_cell = todo_ws.cell(row=todo_row, column=col, value=task["task"])
    task_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    task_cell.border = border_style
    task_cell.font = Font(size=10, bold=True)
    col += 1
    
    # 단계 (체크리스트 기능 통합)
    step_cell = todo_ws.cell(row=todo_row, column=col, value=task.get("step", ""))
    step_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    step_cell.border = border_style
    step_cell.font = Font(size=10)
    col += 1
    
    # 담당자
    person_cell = todo_ws.cell(row=todo_row, column=col, value=task["person"])
    person_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    person_cell.border = border_style
    person_cell.font = Font(size=10)
    col += 1
    
    # 우선순위
    priority_cell = todo_ws.cell(row=todo_row, column=col, value=task["priority"])
    priority_cell.alignment = Alignment(horizontal='center', vertical='center')
    priority_cell.border = border_style
    priority_cell.font = Font(size=10)
    if task["priority"] == "높음":
        priority_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif task["priority"] == "중":
        priority_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        priority_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    col += 1
    
    # 예상소요시간
    duration_cell = todo_ws.cell(row=todo_row, column=col, value=task["duration"])
    duration_cell.alignment = Alignment(horizontal='center', vertical='center')
    duration_cell.border = border_style
    duration_cell.font = Font(size=10)
    col += 1
    
    # 상세내용
    detail_cell = todo_ws.cell(row=todo_row, column=col, value=task["detail"])
    detail_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    detail_cell.border = border_style
    detail_cell.font = Font(size=10)
    col += 1
    
    # 완료여부
    status_cell = todo_ws.cell(row=todo_row, column=col, value=task["status"])
    status_cell.alignment = Alignment(horizontal='center', vertical='center')
    status_cell.border = border_style
    status_cell.font = Font(size=10)
    col += 1
    
    # 비고
    note_cell = todo_ws.cell(row=todo_row, column=col, value=task["note"])
    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    note_cell.border = border_style
    note_cell.font = Font(size=9, italic=True)
    
    todo_row += 1

# 할 일 목록 열 너비 조정 (단계 컬럼 추가로 인한 조정)
todo_column_widths = {'C': 18, 'D': 35, 'E': 15, 'F': 20, 'G': 12, 'H': 15, 'I': 40, 'J': 12, 'K': 30}
for col, width in todo_column_widths.items():
    todo_ws.column_dimensions[col].width = width

# 행 높이 조정
for row in range(2, todo_ws.max_row + 1):
    todo_ws.row_dimensions[row].height = 40
todo_ws.row_dimensions[1].height = 40

# 완료여부 드롭다운 (단계 컬럼 추가로 컬럼 위치 변경: I -> J)
todo_status_validation = DataValidation(type="list", formula1='"미완료,진행중,완료,보류"', allow_blank=True)
todo_ws.add_data_validation(todo_status_validation)
todo_status_validation.add(f'J2:J{todo_ws.max_row}')

# 필터 추가
todo_ws.auto_filter.ref = f'C1:{get_column_letter(todo_ws.max_column)}{todo_ws.max_row}'

# 조건부 서식: 완료여부별 색상 (완료여부 컬럼 J로 변경)
todo_complete_rule = CellIsRule(operator='equal', formula=['"완료"'], 
                               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
todo_ws.conditional_formatting.add(f'J2:J{todo_ws.max_row}', todo_complete_rule)

todo_progress_rule = CellIsRule(operator='equal', formula=['"진행중"'], 
                               fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
todo_ws.conditional_formatting.add(f'J2:J{todo_ws.max_row}', todo_progress_rule)

todo_pending_rule = CellIsRule(operator='equal', formula=['"미완료"'], 
                              fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
todo_ws.conditional_formatting.add(f'J2:J{todo_ws.max_row}', todo_pending_rule)

# 조건부 서식: 우선순위별 행 색상 (높음만, 컬럼 위치 변경: F -> G)
todo_high_priority = CellIsRule(operator='equal', formula=['"높음"'], 
                                fill=PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"))
todo_ws.conditional_formatting.add(f'G2:G{todo_ws.max_row}', todo_high_priority)

# ========== 4. 참고자료 시트 ==========
resources_ws = wb.create_sheet("참고자료")
create_sidebar(resources_ws, "참고자료")

resources_headers = ["카테고리", "자료명", "유형", "위치/링크", "설명", "업데이트일"]
for col_idx, header in enumerate(resources_headers, start=3):
    cell = resources_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

resources_data = [
    {"category": "매뉴얼", "name": "STARNET 설치 매뉴얼", "type": "파일", 
     "location": "C:\\Documents\\Manual\\STARNET.pdf", "desc": "STARNET 설치 및 설정 가이드", 
     "date": "2025-11-15"},
    {"category": "매뉴얼", "name": "서명 템플릿", "type": "파일", 
     "location": "C:\\Documents\\Template\\signature.html", "desc": "이메일 서명 템플릿", 
     "date": "2025-11-15"},
    {"category": "링크", "name": "내부 시스템", "type": "URL", 
     "location": "https://internal.company.com", "desc": "회사 내부 시스템", 
     "date": "2025-11-15"},
]

resources_row = 2
for item in resources_data:
    col = 3
    for key in ["category", "name", "type", "location", "desc", "date"]:
        cell = resources_ws.cell(row=resources_row, column=col, value=item[key])
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if key == "category":
            cell.fill = category_colors.get(item[key], category_colors["기타"])
        col += 1
    resources_row += 1

resources_column_widths = {'C': 15, 'D': 30, 'E': 12, 'F': 40, 'G': 40, 'H': 12}
for col, width in resources_column_widths.items():
    resources_ws.column_dimensions[col].width = width

for row in range(2, resources_ws.max_row + 1):
    resources_ws.row_dimensions[row].height = 30
resources_ws.row_dimensions[1].height = 35

resources_ws.auto_filter.ref = f'C1:{get_column_letter(resources_ws.max_column)}{resources_ws.max_row}'

# ========== 6. 담당자 정보 시트 ==========
contacts_ws = wb.create_sheet("담당자 정보")
create_sidebar(contacts_ws, "담당자 정보")

contacts_headers = ["이름", "직급", "연락처", "이메일", "담당 업무", "출근/재택", "비고"]
for col_idx, header in enumerate(contacts_headers, start=3):
    cell = contacts_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

contacts_data = [
    {"name": "정인님", "position": "신입", "phone": "", "email": "", 
     "work": "인수인계 대상", "schedule": "출근", "note": ""},
    {"name": "영화 과장님", "position": "과장", "phone": "", "email": "", 
     "work": "침퓨즈(APRO, DCC) 담당", "schedule": "재택/출근", "note": ""},
    {"name": "승은 차장님", "position": "차장", "phone": "", "email": "", 
     "work": "악세스(ACCSS) 담당", "schedule": "재택", "note": ""},
    {"name": "현정 과장님", "position": "과장", "phone": "", "email": "", 
     "work": "악세스, 커넥터 EMS 담당", "schedule": "출근", "note": ""},
    {"name": "미경 선배님", "position": "선배", "phone": "", "email": "", 
     "work": "커넥터 EMS, 한국성전, 샘플 담당", "schedule": "출근", "note": ""},
    {"name": "송은", "position": "", "phone": "", "email": "", 
     "work": "샘플 담당", "schedule": "출근", "note": ""},
]

contacts_row = 2
for item in contacts_data:
    col = 3
    for key in ["name", "position", "phone", "email", "work", "schedule", "note"]:
        cell = contacts_ws.cell(row=contacts_row, column=col, value=item[key])
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        col += 1
    contacts_row += 1

contacts_column_widths = {'C': 15, 'D': 12, 'E': 15, 'F': 25, 'G': 30, 'H': 15, 'I': 25}
for col, width in contacts_column_widths.items():
    contacts_ws.column_dimensions[col].width = width

for row in range(2, contacts_ws.max_row + 1):
    contacts_ws.row_dimensions[row].height = 30
contacts_ws.row_dimensions[1].height = 35

contacts_ws.auto_filter.ref = f'C1:{get_column_letter(contacts_ws.max_column)}{contacts_ws.max_row}'

# ========== 7. 고객사/거래처 정보 시트 ==========
clients_ws = wb.create_sheet("고객사거래처")
create_sidebar(clients_ws, "고객사거래처")

clients_headers = ["고객사명", "담당자", "연락처", "주요 업무", "중요 사항", "비고"]
for col_idx, header in enumerate(clients_headers, start=3):
    cell = clients_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

clients_data = [
    {"name": "침퓨즈", "person": "영화 과장님", "contact": "", 
     "work": "APRO, DCC 관련 업무", "important": "인수인계 필수", "note": ""},
    {"name": "악세스(ACCSS)", "person": "승은 차장님, 현정 과장님", "contact": "", 
     "work": "노이즈필터 관련 업무", "important": "정인님 OJT 포함", "note": ""},
    {"name": "커넥터 EMS 3사", "person": "미경 선배님", "contact": "", 
     "work": "EMS 관련 전체 업무", "important": "3개사 전체 인수인계 필요", "note": ""},
    {"name": "한국성전", "person": "미경 선배님", "contact": "", 
     "work": "FCST 관련 업무", "important": "", "note": ""},
    {"name": "샘플", "person": "송은, 미경 선배님", "contact": "", 
     "work": "유상 및 발주 부분", "important": "이레귤러 변경 부분 주의", "note": ""},
]

clients_row = 2
for item in clients_data:
    col = 3
    for key in ["name", "person", "contact", "work", "important", "note"]:
        cell = clients_ws.cell(row=clients_row, column=col, value=item[key])
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        col += 1
    clients_row += 1

clients_column_widths = {'C': 20, 'D': 20, 'E': 15, 'F': 30, 'G': 30, 'H': 25}
for col, width in clients_column_widths.items():
    clients_ws.column_dimensions[col].width = width

for row in range(2, clients_ws.max_row + 1):
    clients_ws.row_dimensions[row].height = 30
clients_ws.row_dimensions[1].height = 35

clients_ws.auto_filter.ref = f'C1:{get_column_letter(clients_ws.max_column)}{clients_ws.max_row}'

# ========== 8. SAP 사용법 시트 ==========
sap_ws = wb.create_sheet("SAP사용법")
create_sidebar(sap_ws, "SAP사용법")

sap_headers = ["기능명", "메뉴경로", "사용목적", "단계별사용법", "스크린샷위치", "주의사항", "트랜잭션코드", "빈도"]
for col_idx, header in enumerate(sap_headers, start=3):
    cell = sap_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

sap_data = [
    {"function": "발주 입력", "menu": "MM > 구매 > 구매오더", "purpose": "고객사 발주 정보 입력", 
     "steps": "1. 구매오더 생성 2. 거래처 선택 3. 품목 입력 4. 수량/단가 입력 5. 저장", 
     "screenshot": "C:\\Documents\\SAP\\발주입력.png", "note": "단가 확인 필수", "tcode": "ME21N", "freq": "자주"},
    {"function": "출하 처리", "menu": "SD > 출하 > 출하 처리", "purpose": "출하 정보 입력 및 처리", 
     "steps": "1. 출하 문서 생성 2. 납품처 선택 3. 품목/수량 입력 4. 저장", 
     "screenshot": "C:\\Documents\\SAP\\출하처리.png", "note": "재고 확인 후 처리", "tcode": "VL01N", "freq": "자주"},
    {"function": "재고 조회", "menu": "MM > 재고 > 재고 조회", "purpose": "현재 재고 수량 확인", 
     "steps": "1. 재고 조회 화면 진입 2. 품목 코드 입력 3. 조회", 
     "screenshot": "C:\\Documents\\SAP\\재고조회.png", "note": "", "tcode": "MMBE", "freq": "자주"},
    {"function": "고객 정보 조회", "menu": "SD > 마스터 데이터 > 고객", "purpose": "고객사 정보 확인", 
     "steps": "1. 고객 번호 입력 2. 조회", 
     "screenshot": "C:\\Documents\\SAP\\고객조회.png", "note": "", "tcode": "XD03", "freq": "가끔"},
]

sap_row = 2
for item in sap_data:
    col = 3
    for key in ["function", "menu", "purpose", "steps", "screenshot", "note", "tcode", "freq"]:
        cell = sap_ws.cell(row=sap_row, column=col, value=item.get(key, ""))
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if key == "freq":
            if item.get(key) == "자주":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif item.get(key) == "가끔":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        col += 1
    sap_row += 1

sap_column_widths = {'C': 20, 'D': 30, 'E': 25, 'F': 40, 'G': 30, 'H': 25, 'I': 15, 'J': 12}
for col, width in sap_column_widths.items():
    sap_ws.column_dimensions[col].width = width

for row in range(2, sap_ws.max_row + 1):
    sap_ws.row_dimensions[row].height = 50
sap_ws.row_dimensions[1].height = 35

sap_ws.auto_filter.ref = f'C1:{get_column_letter(sap_ws.max_column)}{sap_ws.max_row}'

# ========== 13. 시스템/도구 사용법 시트 ==========
system_ws = wb.create_sheet("시스템도구")
create_sidebar(system_ws, "시스템도구")

system_headers = ["시스템도구명", "용도", "접근방법", "주요기능", "사용방법", "단축키팁", "문제해결", "연락처"]
for col_idx, header in enumerate(system_headers, start=3):
    cell = system_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

system_data = [
    {"name": "STARNET", "purpose": "주문 관리 시스템", "access": "프로그램 실행 > 로그인", 
     "features": "주문 조회, 발주 처리, 재고 관리", "usage": "1. 로그인 2. 메뉴 선택 3. 작업 수행", 
     "shortcut": "Ctrl+S: 저장, Ctrl+F: 찾기", "trouble": "접속 안 될 시 IT팀 연락", "contact": "IT팀 내선 1234"},
    {"name": "Outlook", "purpose": "이메일 및 일정 관리", "access": "Office 프로그램 실행", 
     "features": "메일 송수신, 일정 관리, 연락처", "usage": "기본 Outlook 사용법", 
     "shortcut": "Ctrl+N: 새 메일, Ctrl+R: 답장", "trouble": "서명 설정은 도구 > 옵션", "contact": ""},
    {"name": "Excel", "purpose": "데이터 관리 및 분석", "access": "Office 프로그램 실행", 
     "features": "데이터 입력, 계산, 차트", "usage": "템플릿 파일 사용", 
     "shortcut": "Ctrl+C: 복사, Ctrl+V: 붙여넣기, F2: 편집", "trouble": "", "contact": ""},
]

system_row = 2
for item in system_data:
    col = 3
    for key in ["name", "purpose", "access", "features", "usage", "shortcut", "trouble", "contact"]:
        cell = system_ws.cell(row=system_row, column=col, value=item.get(key, ""))
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        col += 1
    system_row += 1

system_column_widths = {'C': 20, 'D': 25, 'E': 20, 'F': 30, 'G': 40, 'H': 25, 'I': 30, 'J': 15}
for col, width in system_column_widths.items():
    system_ws.column_dimensions[col].width = width

for row in range(2, system_ws.max_row + 1):
    system_ws.row_dimensions[row].height = 40
system_ws.row_dimensions[1].height = 35

system_ws.auto_filter.ref = f'C1:{get_column_letter(system_ws.max_column)}{system_ws.max_row}'

# ========== 14. 용어집/약어 사전 시트 ==========
glossary_ws = wb.create_sheet("용어집")
create_sidebar(glossary_ws, "용어집")

glossary_headers = ["용어약어", "풀네임", "의미설명", "사용예시", "관련업무", "카테고리"]
for col_idx, header in enumerate(glossary_headers, start=3):
    cell = glossary_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

glossary_data = [
    {"term": "APRO", "fullname": "", "meaning": "침퓨즈 관련 업무 프로세스", "example": "APRO 발주 처리", 
     "work": "침퓨즈 인수인계", "category": "고객사"},
    {"term": "DCC", "fullname": "", "meaning": "침퓨즈 관련 업무 프로세스", "example": "DCC 발주 처리", 
     "work": "침퓨즈 인수인계", "category": "고객사"},
    {"term": "FCST", "fullname": "Forecast", "meaning": "수요 예측", "example": "한국성전 FCST 인계", 
     "work": "한국성전 업무", "category": "업무"},
    {"term": "ACCSS", "fullname": "Access", "meaning": "악세스 노이즈필터 관련", "example": "ACCSS 인수인계", 
     "work": "악세스 업무", "category": "고객사"},
    {"term": "EMS", "fullname": "Electronic Manufacturing Services", "meaning": "전자제조 서비스", "example": "커넥터 EMS 3사", 
     "work": "커넥터 업무", "category": "고객사"},
    {"term": "OJT", "fullname": "On-the-Job Training", "meaning": "현장 실무 교육", "example": "정인님 OJT", 
     "work": "인수인계", "category": "교육"},
]

glossary_row = 2
for item in glossary_data:
    col = 3
    for key in ["term", "fullname", "meaning", "example", "work", "category"]:
        cell = glossary_ws.cell(row=glossary_row, column=col, value=item.get(key, ""))
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if key == "term":
            cell.font = Font(bold=True, size=10)
        col += 1
    glossary_row += 1

glossary_column_widths = {'C': 15, 'D': 25, 'E': 35, 'F': 30, 'G': 25, 'H': 15}
for col, width in glossary_column_widths.items():
    glossary_ws.column_dimensions[col].width = width

for row in range(2, glossary_ws.max_row + 1):
    glossary_ws.row_dimensions[row].height = 40
glossary_ws.row_dimensions[1].height = 35

glossary_ws.auto_filter.ref = f'C1:{get_column_letter(glossary_ws.max_column)}{glossary_ws.max_row}'

# ========== 16. 고객처별 상세 가이드 시트 ==========
guide_ws = wb.create_sheet("고객처별가이드")
create_sidebar(guide_ws, "고객처별가이드")

# 제목 행
title_cell = guide_ws.cell(row=1, column=3, value="고객처별 상세 업무 가이드")
title_cell.font = Font(bold=True, size=16, color="366092")
title_cell.alignment = Alignment(horizontal='left', vertical='center')

# 고객처 선택 드롭다운
guide_ws.cell(row=2, column=3, value="고객처 선택:").font = Font(bold=True, size=11)
customer_validation = DataValidation(type="list", formula1='"Parts_일반발주,침퓨즈(APRO/DCC),악세스(ACCSS),커넥터EMS,한국성전(FCST),샘플"', allow_blank=True)
guide_ws.add_data_validation(customer_validation)
customer_cell = guide_ws.cell(row=2, column=4, value="Parts_일반발주")
customer_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
customer_validation.add("D2")

# 헤더 (이미지 구조 참고)
guide_headers = ["단계", "SAP t-code", "method", "설명"]
header_start_row = 4
for col_idx, header in enumerate(guide_headers, start=3):
    cell = guide_ws.cell(row=header_start_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# Parts_일반발주 상세 가이드 데이터 (이미지 참고)
parts_guide_data = [
    {
        "step": "1. CPO 입수",
        "tcode": "",
        "method": "이메일",
        "description": """*주 : 4번 단계 이후 CPO 수정을 요청한 경우, 수정 CPO 입수 하고 발주 진행하기
*주 : 발주 진행 후 대체품번으로 변경 등으로 인해 주문 품번 취소를 해야할 경우, 고객사에 설명 및 확인 회신 받은 후, 사업부에 취소 진행 및 사업부 취소 완료 확인 필수, 그 후 수정 CPO 입수 후 프로세스대로 발주 진행
*발주서 보관 폴더 경로 :"""
    },
    {
        "step": "2. 발주서 변환파일",
        "tcode": "",
        "method": "엑셀파일",
        "description": """1. CPO 내용 복붙
2. GLICS 탭에 GLICS 조회결과 복붙 (*GLICS 1회 최대 15개 품번 검색 가능)
3. 화학물질 열에 결과값(!)이 나오는 품번은 별도 FE (=별도 SAP (PO)로 발주 진행하기
4. 그 외 확인 내용: MOQ, 매입가, 셀링가의 일치여부, 지누텍 : 모터품번의 프로핏 센터(상세내용 하단에 적기)
5. 벤더별 드롭다운 선택 : 일본: JP/ 일본 긴급발주: EO / 싱가폴 일반 or 긴급발주 : SG"""
    },
    {
        "step": "3. Glics 보충설명 확인 등",
        "tcode": "",
        "method": "GLICS",
        "description": """*주 : 1회 최대 조회 가능 품번 수: 15개
*매입가, 대체품번, 보충설명, MOQ 등 확인"""
    },
    {
        "step": "4. GLICS 보충설명 확인 요청 메일 전송",
        "tcode": "",
        "method": "",
        "description": """1. 지누텍만 전송, 메일 내용 : 보충설명 내용 확인, 대체품번, MOQ 등 체크 후 안내 및 수정 CPO 송부 요청
2. DJK는 보충설명 확인 후 CPO 보내심"""
    },
    {
        "step": "5. Segmentation 확인",
        "tcode": "ZMMM_MCS5N",
        "method": "",
        "description": """*MM03 에서 SEGMENTATION 품번 검색해서 품번+SEGMENTATION 없으면 등록이 필요함
*품번이 GCS 품번이어야 함
*주: 업로드 엑셀 양식에서 주의할 부분
① Purchasing group: P01
② Transportation group: 0001
③ Loading group: 0001"""
    },
    {
        "step": "6. eMaster 상신",
        "tcode": "",
        "method": "서비스나우",
        "description": """서비스나우에 업로드 양식 내용 반영 후, 품번 클릭해서 세부 화면에서도 한 번 더 확인하기.
만약 0001이 아니라 1로 나와있으면 수정하기 & 첨부파일도 수정해서 다시 첨부하기."""
    },
    {
        "step": "7. eMaster 승인",
        "tcode": "",
        "method": "서비스나우",
        "description": "승인 완료 대기"
    },
    {
        "step": "8. 품번 등록 (PIR 등록)",
        "tcode": "ME11 (등록), ME13 (디스플레이)",
        "method": "",
        "description": """*벤더코드 주의 (일본: 20015, 일본 긴급발주: 21876, 싱가폴 : 28005)
*Plant. (7521)은 입력 X
*입력 방법은 우측 스크린샷 참고"""
    }
]

# 데이터 입력
guide_row = header_start_row + 1
for item in parts_guide_data:
    col = 3
    for key in ["step", "tcode", "method", "description"]:
        cell = guide_ws.cell(row=guide_row, column=col, value=item.get(key, ""))
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if key == "step":
            cell.font = Font(bold=True, size=10)
        elif key == "description":
            cell.font = Font(size=9)
        col += 1
    guide_row += 1

# 열 너비 조정
guide_column_widths = {'C': 25, 'D': 15, 'E': 15, 'F': 80}
for col, width in guide_column_widths.items():
    guide_ws.column_dimensions[col].width = width

# 행 높이 조정 (설명이 길어서)
for row in range(header_start_row + 1, guide_ws.max_row + 1):
    guide_ws.row_dimensions[row].height = 80
guide_ws.row_dimensions[header_start_row].height = 35

# 필터 추가
guide_ws.auto_filter.ref = f'C{header_start_row}:{get_column_letter(guide_ws.max_column)}{guide_ws.max_row}'

# 안내 문구 추가
info_row = guide_ws.max_row + 2
info_cell = guide_ws.cell(row=info_row, column=3, 
                          value="※ 각 고객처별 상세 가이드는 위 드롭다운에서 선택하거나, 별도 시트로 추가할 수 있습니다.")
info_cell.font = Font(size=9, italic=True, color="666666")
info_cell.alignment = Alignment(horizontal='left', vertical='center')

# ========== 일반 업무 프로세스 섹션 추가 (업무프로세스 시트 통합) ==========
general_process_row = info_row + 3
general_title = guide_ws.cell(row=general_process_row, column=3, value="일반 업무 프로세스")
general_title.font = Font(bold=True, size=14, color="366092")
general_title.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
general_title.alignment = Alignment(horizontal='left', vertical='center')
general_title.border = border_style
guide_ws.merge_cells(f'C{general_process_row}:F{general_process_row}')
general_process_row += 1

# 일반 프로세스 헤더
general_headers = ["업무명", "카테고리", "1단계", "2단계", "3단계", "4단계", "5단계", "주의사항", "관련시스템", "참고문서", "담당자"]
for col_idx, header in enumerate(general_headers, start=3):
    cell = guide_ws.cell(row=general_process_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style
general_process_row += 1

# 일반 프로세스 데이터 (업무프로세스 시트에서 통합)
general_process_data = [
    {"name": "발주 프로세스", "category": "발주", "step1": "고객사 발주 요청 확인", "step2": "SAP에서 발주 입력", 
     "step3": "승인 프로세스 진행", "step4": "발주서 출력 및 전송", "step5": "완료 확인", 
     "note": "긴급 발주 시 별도 프로세스", "system": "SAP", "doc": "발주 매뉴얼", "person": "담당자"},
    {"name": "출하 프로세스", "category": "출하", "step1": "출하 예정 확인", "step2": "출하 정보 입력", 
     "step3": "출하서 작성", "step4": "물류팀 전달", "step5": "출하 완료 처리", 
     "note": "출하 전 재고 확인 필수", "system": "SAP", "doc": "출하 매뉴얼", "person": "담당자"},
    {"name": "샘플 처리", "category": "샘플", "step1": "샘플 요청 확인", "step2": "유상/무상 구분", 
     "step3": "샘플 발주 처리", "step4": "이레귤러 확인", "step5": "완료", 
     "note": "이레귤러 변경 부분 주의", "system": "SAP", "doc": "샘플 매뉴얼", "person": "송은, 미경"},
]

for item in general_process_data:
    col = 3
    for key in ["name", "category", "step1", "step2", "step3", "step4", "step5", "note", "system", "doc", "person"]:
        cell = guide_ws.cell(row=general_process_row, column=col, value=item.get(key, ""))
        cell.border = border_style
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if key == "category":
            cell.fill = category_colors.get(item.get(key, ""), category_colors["기타"])
        col += 1
    general_process_row += 1

# 일반 프로세스 행 높이 조정
for row in range(general_process_row - len(general_process_data) - 1, general_process_row):
    guide_ws.row_dimensions[row].height = 40

# ========== 17. 워크플로우 시트 생성 ==========
workflow_ws = wb.create_sheet("워크플로우")
create_sidebar(workflow_ws, "워크플로우")

# 제목
workflow_title = workflow_ws.cell(row=1, column=3, value="인수인계 업무 워크플로우")
workflow_title.font = Font(bold=True, size=18, color="366092")
workflow_title.alignment = Alignment(horizontal='left', vertical='center')

# 설명
workflow_desc = workflow_ws.cell(row=2, column=3, value="인수인계 업무의 전체 흐름과 각 단계별 연결된 시트를 확인하세요")
workflow_desc.font = Font(size=11, italic=True, color="666666")
workflow_desc.alignment = Alignment(horizontal='left', vertical='center')

row = 4

# 워크플로우 시나리오들
workflows = [
    {
        "title": "일일 업무 시작",
        "description": "매일 아침 인수인계 업무를 시작할 때",
        "steps": [
            {"step": "1. 대시보드 확인", "action": "오늘의 일정과 진행 상황 확인", "link": "대시보드"},
            {"step": "2. 인수인계 스케줄 확인", "action": "오늘 시간대별 일정 확인", "link": "인수인계 스케줄"},
            {"step": "3. 할 일 목록 확인", "action": "오늘 해야 할 작업 확인", "link": "할 일 목록"}
        ]
    },
    {
        "title": "고객처별 발주 업무",
        "description": "특정 고객처의 발주 업무를 수행할 때",
        "steps": [
            {"step": "1. 고객처별 가이드 선택", "action": "해당 고객처 가이드 확인", "link": "고객처별가이드"},
            {"step": "2. 고객처별 가이드 확인", "action": "고객처별 상세 프로세스 확인 (일반 프로세스 포함)", "link": "고객처별가이드"},
            {"step": "3. SAP 사용법 확인", "action": "필요한 SAP t-code 확인", "link": "SAP사용법"},
            {"step": "4. 단계별 수행", "action": "가이드에 따라 단계별 수행", "link": "고객처별가이드"},
            {"step": "5. 할 일 목록 업데이트", "action": "각 단계별 완료 여부 업데이트", "link": "할 일 목록"}
        ]
    },
    {
        "title": "문제 발생 시 해결",
        "description": "업무 중 문제가 발생했을 때",
        "steps": [
            {"step": "1. 엑셀 검색 기능 활용", "action": "Ctrl+F로 키워드 검색", "link": ""},
            {"step": "2. 용어집 확인", "action": "모르는 용어 확인", "link": "용어집"},
            {"step": "3. 관련 시트 확인", "action": "고객처별가이드, SAP사용법 등 관련 시트 확인", "link": ""},
            {"step": "4. 담당자 문의", "action": "해결 안 되면 담당자에게 문의", "link": "담당자 정보"}
        ]
    },
    {
        "title": "전체 인수인계 추적",
        "description": "인수인계 전체 진행 상황을 추적할 때",
        "steps": [
            {"step": "1. 할 일 목록 필터", "action": "미완료 항목만 필터링 (단계별 확인 가능)", "link": "할 일 목록"},
            {"step": "2. 스케줄 확인", "action": "일정별 진행 상황 확인", "link": "인수인계 스케줄"},
            {"step": "3. 워크플로우 확인", "action": "전체 업무 흐름 파악", "link": "워크플로우"}
        ]
    }
]

for workflow in workflows:
    # 워크플로우 제목
    title_cell = workflow_ws.cell(row=row, column=3, value=workflow["title"])
    title_cell.font = Font(bold=True, size=14, color="366092")
    title_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    title_cell.border = border_style
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    workflow_ws.row_dimensions[row].height = 25
    row += 1
    
    # 설명
    desc_cell = workflow_ws.cell(row=row, column=3, value=workflow["description"])
    desc_cell.font = Font(size=10, italic=True, color="666666")
    desc_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # 단계별 플로우
    for step_info in workflow["steps"]:
        # 단계 번호와 설명
        step_text = f"{step_info['step']}: {step_info['action']}"
        step_cell = workflow_ws.cell(row=row, column=3, value=step_text)
        step_cell.font = Font(size=10)
        step_cell.border = border_style
        step_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        
        # 링크 버튼
        if step_info['link']:
            link_cell = workflow_ws.cell(row=row, column=4, value=f"→ {step_info['link']}")
            link_cell.font = Font(size=9, color="0000FF", underline="single")
            link_cell.border = border_style
            link_cell.alignment = Alignment(horizontal='center', vertical='center')
            link_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            # 하이퍼링크 추가
            for sheet_info in sheets_info:
                if step_info['link'] in sheet_info["name"]:
                    link_cell.hyperlink = f"#'{sheet_info['name']}'!A1"
                    break
        
        row += 1
    
    # 워크플로우 간 구분
    row += 1

# 열 너비 조정
workflow_ws.column_dimensions['C'].width = 50
workflow_ws.column_dimensions['D'].width = 25

# 각 시트에 연관 시트 링크 추가 함수
def add_related_links(ws, related_sheets, row_start=1):
    """각 시트 하단에 연관 시트 링크 추가"""
    max_row = ws.max_row
    link_row = max_row + 3
    
    if related_sheets:
        related_title = ws.cell(row=link_row, column=3, value="관련 시트")
        related_title.font = Font(bold=True, size=11, color="366092")
        related_title.alignment = Alignment(horizontal='left', vertical='center')
        link_row += 1
        
        for i, sheet_name in enumerate(related_sheets):
            col = 3 + (i % 3)
            r = link_row + (i // 3)
            
            link_cell = ws.cell(row=r, column=col, value=f"→ {sheet_name}")
            link_cell.font = Font(size=9, color="0000FF", underline="single")
            link_cell.border = border_style
            link_cell.alignment = Alignment(horizontal='left', vertical='center')
            link_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # 하이퍼링크 추가
            for sheet_info in sheets_info:
                if sheet_name in sheet_info["name"]:
                    link_cell.hyperlink = f"#'{sheet_info['name']}'!A1"
                    break

# 각 시트에 연관 링크 추가
# 고객처별 가이드에 연관 링크 (업무프로세스 통합됨)
add_related_links(guide_ws, ["SAP사용법", "담당자 정보", "고객사거래처"])

# SAP 사용법에 연관 링크
add_related_links(sap_ws, ["고객처별가이드", "시스템도구"])

# 할 일 목록에 연관 링크 (체크리스트 기능 통합됨)
add_related_links(todo_ws, ["인수인계 스케줄", "워크플로우"])


# ========== 18. 구현 가이드 시트 생성 ==========
guide_impl_ws = wb.create_sheet("구현가이드")
create_sidebar(guide_impl_ws, "구현가이드")

# 제목
title_cell = guide_impl_ws.cell(row=1, column=3, value="엑셀 기능 구현 가이드")
title_cell.font = Font(bold=True, size=18, color="366092")
title_cell.alignment = Alignment(horizontal='left', vertical='center')

subtitle_cell = guide_impl_ws.cell(row=2, column=3, value="이 문서에 사용된 모든 엑셀 기능의 구현 방법을 정리했습니다.")
subtitle_cell.font = Font(size=11, italic=True, color="666666")
subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')

row = 4

# 구현 가이드 데이터
implementation_guides = [
    {
        "category": "1. 기본 설정",
        "items": [
            {
                "feature": "시트 생성 및 이름 변경",
                "python": "ws = wb.create_sheet('시트명')\nws.title = '새 시트명'",
                "excel": "시트 탭 우클릭 → 이름 바꾸기",
                "note": "시트 이름에 특수문자(/, \\, ?, *, []) 사용 불가"
            },
            {
                "feature": "열 너비 조정",
                "python": "ws.column_dimensions['A'].width = 25",
                "excel": "열 경계선 드래그 또는 열 선택 → 우클릭 → 열 너비",
                "note": "숫자는 문자 수 기준"
            },
            {
                "feature": "행 높이 조정",
                "python": "ws.row_dimensions[1].height = 30",
                "excel": "행 경계선 드래그 또는 행 선택 → 우클릭 → 행 높이",
                "note": "숫자는 포인트 단위"
            },
            {
                "feature": "헤더 스타일 설정",
                "python": "cell.fill = PatternFill(start_color='366092', ...)\ncell.font = Font(bold=True, color='FFFFFF')",
                "excel": "셀 선택 → 홈 → 글꼴/채우기 색 → 굵게",
                "note": "색상 코드는 16진수 (예: 366092 = 진한 파랑)"
            }
        ]
    },
    {
        "category": "2. 조건부 서식",
        "items": [
            {
                "feature": "완료여부별 색상 자동 변경",
                "python": "from openpyxl.formatting.rule import CellIsRule\nrule = CellIsRule(operator='equal', formula=['\"완료\"'],\n                  fill=PatternFill(start_color='C6EFCE', ...))\nws.conditional_formatting.add('I2:I100', rule)",
                "excel": "셀 범위 선택 → 홈 → 조건부 서식 → 셀 값 기준 → 같음 → '완료' → 서식 → 채우기 색",
                "note": "완료=초록(C6EFCE), 진행중=노랑(FFEB9C), 미완료=빨강(FFC7CE)"
            },
            {
                "feature": "오늘 날짜 하이라이트",
                "python": "from openpyxl.formatting.rule import FormulaRule\nrule = FormulaRule(formula=['$C2=TODAY()'],\n                  fill=PatternFill(start_color='FFD966', ...))\nws.conditional_formatting.add('C2:M100', rule)",
                "excel": "셀 범위 선택 → 조건부 서식 → 수식 사용 → =$C2=TODAY() → 서식 설정",
                "note": "$C2는 절대 열 참조, TODAY()는 오늘 날짜 반환"
            },
            {
                "feature": "우선순위별 강조",
                "python": "rule = CellIsRule(operator='equal', formula=['\"높음\"'],\n                  fill=PatternFill(start_color='FFE6E6', ...))",
                "excel": "조건부 서식 → 셀 값 기준 → 같음 → '높음'",
                "note": "높은 우선순위 항목 자동 강조"
            }
        ]
    },
    {
        "category": "3. 수식 및 함수",
        "items": [
            {
                "feature": "COUNTIF - 진행률 자동 계산",
                "python": "cell.value = '=COUNTIF(\\'할 일 목록\\'!I:I, \"완료\")'",
                "excel": "=COUNTIF(할 일 목록!I:I, \"완료\")",
                "note": "I열에서 '완료'인 셀 개수 자동 계산"
            },
            {
                "feature": "TODAY - 오늘 날짜 자동 표시",
                "python": "cell.value = '=TODAY()'\ncell.number_format = 'yyyy-mm-dd'",
                "excel": "=TODAY()",
                "note": "매일 자동으로 오늘 날짜 업데이트"
            },
            {
                "feature": "SUM - 총계 계산",
                "python": "cell.value = '=SUM(A1:A10)'",
                "excel": "=SUM(A1:A10) 또는 =SUM(A1:A10)",
                "note": "범위 내 숫자 합계"
            },
            {
                "feature": "IF - 조건부 계산",
                "python": "cell.value = '=IF(총계>0, 완료/총계*100, 0)'",
                "excel": "=IF(조건, 참일때값, 거짓일때값)",
                "note": "완료율 계산 등에 사용"
            },
            {
                "feature": "시트 간 참조",
                "python": "cell.value = '=COUNTIF(\\'시트명\\'!A:A, 조건)'",
                "excel": "='시트명'!A1 또는 =시트명!A1",
                "note": "시트명에 공백/특수문자 있으면 작은따옴표 필수"
            }
        ]
    },
    {
        "category": "4. 데이터 유효성 검사",
        "items": [
            {
                "feature": "드롭다운 목록 생성",
                "python": "from openpyxl.worksheet.datavalidation import DataValidation\nvalidation = DataValidation(type='list',\n                              formula1='\"미완료,진행중,완료,보류\"')\nws.add_data_validation(validation)\nvalidation.add('I2:I100')",
                "excel": "셀 선택 → 데이터 → 데이터 유효성 검사 → 목록 → 원본에 값 입력",
                "note": "콤마로 구분된 목록 또는 셀 범위 참조"
            },
            {
                "feature": "오류 메시지 설정",
                "python": "validation.error = '목록에서 선택해주세요'\nvalidation.errorTitle = '잘못된 입력'",
                "excel": "데이터 유효성 검사 → 오류 경고 탭 → 메시지 입력",
                "note": "잘못된 값 입력 시 표시할 메시지"
            }
        ]
    },
    {
        "category": "5. 하이퍼링크",
        "items": [
            {
                "feature": "시트 간 하이퍼링크",
                "python": "cell.hyperlink = \"#'시트명'!A1\"\ncell.font = Font(color='0000FF', underline='single')",
                "excel": "셀 선택 → 삽입 → 하이퍼링크 → 이 문서의 위치 → 시트 선택",
                "note": "클릭하면 해당 시트로 이동"
            },
            {
                "feature": "하이퍼링크 스타일",
                "python": "cell.font = Font(color='0000FF', underline='single')",
                "excel": "하이퍼링크 셀 → 홈 → 글꼴 색 → 파란색",
                "note": "기본적으로 파란색 밑줄"
            }
        ]
    },
    {
        "category": "6. 필터 및 정렬",
        "items": [
            {
                "feature": "자동 필터 추가",
                "python": "ws.auto_filter.ref = 'A1:Z100'",
                "excel": "데이터 범위 선택 → 데이터 → 필터",
                "note": "헤더 행에 드롭다운 화살표 표시"
            },
            {
                "feature": "다중 조건 필터링",
                "python": "자동 필터 사용 시 엑셀에서 직접 설정",
                "excel": "필터 드롭다운 → 여러 조건 선택",
                "note": "여러 열에 동시에 필터 적용 가능"
            }
        ]
    },
    {
        "category": "7. 사이드바 네비게이션",
        "items": [
            {
                "feature": "사이드바 열 생성",
                "python": "ws.column_dimensions['A'].width = 28\n# A열에 배경색 적용",
                "excel": "A열 선택 → 열 너비 조정 → 배경색 적용",
                "note": "A열을 사이드바로 사용"
            },
            {
                "feature": "그룹 헤더 추가",
                "python": "cell.fill = PatternFill(start_color='DEEBF7', ...)\ncell.font = Font(bold=True)",
                "excel": "셀 선택 → 배경색 + 굵게",
                "note": "그룹별 색상으로 구분"
            },
            {
                "feature": "현재 위치 강조",
                "python": "if sheet_name == current_sheet_name:\n    cell.fill = PatternFill(start_color='FFD966', ...)\n    cell.border = Border(left=Side(style='thick', ...))",
                "excel": "조건부 서식 사용 → 수식: =셀주소=\"현재시트명\"",
                "note": "노란색 배경 + 굵은 왼쪽 테두리"
            }
        ]
    },
    {
        "category": "8. 셀 병합 및 서식",
        "items": [
            {
                "feature": "셀 병합",
                "python": "ws.merge_cells('A1:A10')",
                "excel": "셀 범위 선택 → 홈 → 병합하고 가운데 맞춤",
                "note": "날짜/요일 같은 경우 여러 행 병합"
            },
            {
                "feature": "텍스트 줄바꿈",
                "python": "cell.alignment = Alignment(wrap_text=True)",
                "excel": "셀 선택 → 홈 → 텍스트 줄 바꿈",
                "note": "긴 텍스트 자동 줄바꿈"
            },
            {
                "feature": "숫자 형식 지정",
                "python": "cell.number_format = '0.0\"%\"'  # 백분율\ncell.number_format = 'yyyy-mm-dd'  # 날짜",
                "excel": "셀 선택 → 우클릭 → 셀 서식 → 숫자 탭",
                "note": "백분율, 날짜, 통화 등 형식 지정"
            }
        ]
    },
    {
        "category": "9. 색상 코드",
        "items": [
            {
                "feature": "주요 색상 코드",
                "python": "366092 = 진한 파랑 (헤더)\nC6EFCE = 연한 초록 (완료)\nFFEB9C = 연한 노랑 (진행중)\nFFC7CE = 연한 빨강 (미완료)\nFFD966 = 노랑 (오늘/강조)",
                "excel": "홈 → 채우기 색 → 기타 색 → 사용자 지정 → RGB 입력",
                "note": "16진수 색상 코드 사용 (RRGGBB)"
            }
        ]
    },
    {
        "category": "10. 워크플로우 시트",
        "items": [
            {
                "feature": "시나리오별 플로우 작성",
                "python": "단계별로 행 추가 → 각 단계에 링크 버튼 추가",
                "excel": "수동으로 단계 입력 → 하이퍼링크 추가",
                "note": "업무 흐름을 시각적으로 표현"
            },
            {
                "feature": "링크 버튼 생성",
                "python": "cell.hyperlink = \"#'시트명'!A1\"\ncell.fill = PatternFill(start_color='E2EFDA', ...)",
                "excel": "셀에 텍스트 입력 → 하이퍼링크 추가 → 배경색 적용",
                "note": "클릭 가능한 버튼처럼 보이게 스타일링"
            }
        ]
    }
]

# 구현 가이드 작성
for guide_section in implementation_guides:
    # 카테고리 제목
    cat_cell = guide_impl_ws.cell(row=row, column=3, value=guide_section["category"])
    cat_cell.font = Font(bold=True, size=14, color="366092")
    cat_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cat_cell.border = border_style
    cat_cell.alignment = Alignment(horizontal='left', vertical='center')
    guide_impl_ws.row_dimensions[row].height = 25
    row += 1
    
    # 헤더
    headers = ["기능", "Python 코드", "엑셀에서 하는 방법", "주의사항"]
    for col_idx, header in enumerate(headers, start=3):
        cell = guide_impl_ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_style
    guide_impl_ws.row_dimensions[row].height = 30
    row += 1
    
    # 각 기능별 설명
    for item in guide_section["items"]:
        # 기능명
        feature_cell = guide_impl_ws.cell(row=row, column=3, value=item["feature"])
        feature_cell.font = Font(bold=True, size=10)
        feature_cell.border = border_style
        feature_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Python 코드
        python_cell = guide_impl_ws.cell(row=row, column=4, value=item["python"])
        python_cell.font = Font(size=9, name="Courier New")
        python_cell.border = border_style
        python_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        python_cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
        
        # 엑셀에서 하는 방법
        excel_cell = guide_impl_ws.cell(row=row, column=5, value=item["excel"])
        excel_cell.font = Font(size=9)
        excel_cell.border = border_style
        excel_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 주의사항
        note_cell = guide_impl_ws.cell(row=row, column=6, value=item["note"])
        note_cell.font = Font(size=9, italic=True, color="666666")
        note_cell.border = border_style
        note_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        guide_impl_ws.row_dimensions[row].height = 60
        row += 1
    
    # 섹션 간 구분
    row += 1

# 열 너비 조정
guide_impl_ws.column_dimensions['C'].width = 25
guide_impl_ws.column_dimensions['D'].width = 50
guide_impl_ws.column_dimensions['E'].width = 40
guide_impl_ws.column_dimensions['F'].width = 30

# 필터 추가
guide_impl_ws.auto_filter.ref = f'C4:{get_column_letter(guide_impl_ws.max_column)}{guide_impl_ws.max_row}'

# ========== 파일 저장 ==========
try:
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
    desktop_path = winreg.QueryValueEx(key, "Desktop")[0]
    winreg.CloseKey(key)
except:
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    if not os.path.exists(desktop_path):
        desktop_path = os.path.join(os.path.expanduser('~'), 'Documents')

filename = os.path.join(desktop_path, "인수인계_Notion스타일_v8_최종최적화.xlsx")
wb.save(filename)
print(f"Notion 스타일 인수인계 문서가 바탕화면에 생성되었습니다: {filename}")
print(f"\n포함된 시트 ({len(wb.sheetnames)}개):")
try:
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet_name}")
except:
    print(f"  총 {len(wb.sheetnames)}개의 시트가 생성되었습니다.")
print(f"\n주요 기능:")
print(f"  - 각 시트에 사이드바 네비게이션 (그룹핑)")
print(f"  - 오늘 날짜 자동 표시 및 일정 개수 계산")
print(f"  - 조건부 서식: 완료여부별 자동 색상 변경")
print(f"  - 조건부 서식: 오늘 일정 자동 하이라이트")
print(f"  - 조건부 서식: 우선순위별 자동 강조")
print(f"  - 시트 간 하이퍼링크로 빠른 이동")
print(f"  - 필터 및 정렬 기능")
print(f"  - 워크플로우 시트로 업무 흐름 파악")

