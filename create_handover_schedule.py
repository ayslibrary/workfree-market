import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
import os
import winreg

# 날짜 범위 설정
start_date = datetime(2025, 11, 17)
end_date = datetime(2025, 11, 28)

# 엑셀 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "인수인계 스케줄"

# 헤더 스타일 설정
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 헤더 작성
headers = ["날짜", "요일", "시간", "주제/내용", "담당자", "비고"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style

# 날짜별로 행 생성
row = 2
current_date = start_date
weekdays = ['월', '화', '수', '목', '금', '토', '일']

while current_date <= end_date:
    weekday = weekdays[current_date.weekday()]
    
    # 날짜 행 (병합)
    date_cell = ws.cell(row=row, column=1, value=current_date.strftime("%Y-%m-%d"))
    date_cell.alignment = Alignment(horizontal='center', vertical='center')
    date_cell.border = border_style
    date_cell.font = Font(bold=True, size=10)
    
    weekday_cell = ws.cell(row=row, column=2, value=weekday)
    weekday_cell.alignment = Alignment(horizontal='center', vertical='center')
    weekday_cell.border = border_style
    
    # 시간대별 행 생성 (9시~18시, 1시간 단위)
    time_slots = []
    for hour in range(9, 18):
        time_str = f"{hour:02d}:00 - {hour+1:02d}:00"
        time_slots.append(time_str)
    
    # 첫 번째 시간대만 날짜와 요일 표시
    for time_idx, time_slot in enumerate(time_slots):
        if time_idx == 0:
            # 날짜와 요일은 첫 번째 행에만
            pass
        else:
            # 나머지 시간대는 날짜/요일 셀 병합을 위해 빈 값
            ws.cell(row=row + time_idx, column=1, value="").border = border_style
            ws.cell(row=row + time_idx, column=2, value="").border = border_style
        
        # 시간
        time_cell = ws.cell(row=row + time_idx, column=3, value=time_slot)
        time_cell.alignment = Alignment(horizontal='center', vertical='center')
        time_cell.border = border_style
        
        # 주제/내용 (빈 셀)
        content_cell = ws.cell(row=row + time_idx, column=4, value="")
        content_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        content_cell.border = border_style
        
        # 담당자 (빈 셀)
        person_cell = ws.cell(row=row + time_idx, column=5, value="")
        person_cell.alignment = Alignment(horizontal='center', vertical='center')
        person_cell.border = border_style
        
        # 비고 (빈 셀)
        note_cell = ws.cell(row=row + time_idx, column=6, value="")
        note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        note_cell.border = border_style
    
    # 날짜와 요일 셀 병합
    if len(time_slots) > 1:
        ws.merge_cells(f'A{row}:A{row + len(time_slots) - 1}')
        ws.merge_cells(f'B{row}:B{row + len(time_slots) - 1}')
    
    row += len(time_slots)
    current_date += timedelta(days=1)

# 열 너비 조정
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 30

# 행 높이 조정
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 25

# 첫 번째 행 높이 조정
ws.row_dimensions[1].height = 30

# 파일 저장 (바탕화면에 저장)
# Windows 바탕화면 경로 가져오기
try:
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
    desktop_path = winreg.QueryValueEx(key, "Desktop")[0]
    winreg.CloseKey(key)
except:
    # 실패 시 기본 경로 사용
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    if not os.path.exists(desktop_path):
        desktop_path = os.path.join(os.path.expanduser('~'), 'Documents')

filename = os.path.join(desktop_path, "신입_인수인계_스케줄표.xlsx")
wb.save(filename)
print(f"엑셀 파일이 바탕화면에 생성되었습니다: {filename}")

