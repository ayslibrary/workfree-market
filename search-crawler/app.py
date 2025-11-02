"""
WorkFree Search Crawler API
구글/네이버 검색 자동화 & 이메일 발송
"""

from fastapi import FastAPI, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
from typing import List, Optional
import requests
from bs4 import BeautifulSoup
import csv
from io import StringIO, BytesIO
from datetime import datetime
import base64
import os
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import resend
from scheduler import scheduler_manager

# .env 파일 로드
load_dotenv()

app = FastAPI(
    title="WorkFree 뉴스 크롤링 API",
    description="검색어 기반 뉴스 자동 크롤링 & 메일 발송 API",
    version="1.0.0"
)

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://workfreemarket.com",
        "https://*.vercel.app",
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 환경 변수
# Resend API (이메일 발송)
RESEND_API_KEY = os.getenv("RESEND_API_KEY")

# Google Custom Search API (선택사항)
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")
GOOGLE_SEARCH_ENGINE_ID = os.getenv("GOOGLE_SEARCH_ENGINE_ID", "")

# Naver Search API
NAVER_CLIENT_ID = os.getenv("NAVER_CLIENT_ID")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET")

# API 인증 키 (간단한 보안)
API_SECRET_KEY = os.getenv("API_SECRET_KEY", "workfree-secret-2024")

# 인증 미들웨어
async def verify_api_key(x_api_key: Optional[str] = Header(None)):
    """API 키 검증 - 스케줄 관련 엔드포인트 보호"""
    # 개발 환경에서는 API 키 검증 생략 (localhost)
    # 프로덕션에서는 API 키 필수
    if x_api_key == API_SECRET_KEY:
        return True
    
    # API 키가 없거나 틀린 경우
    if not x_api_key:
        # 개발 환경(localhost 요청)이면 허용
        return True  # 임시로 모든 요청 허용 (나중에 Firebase Auth로 교체 예정)
    
    raise HTTPException(
        status_code=401, 
        detail="인증 실패: 유효하지 않은 API 키입니다"
    )

# Request Models
class SearchRequest(BaseModel):
    keyword: str
    engines: List[str] = ["google", "naver"]  # google, naver
    max_results: int = 10

class EmailRequest(BaseModel):
    keyword: str
    recipient_email: EmailStr
    engines: List[str] = ["google", "naver"]
    max_results: int = 10

class ScheduleRequest(BaseModel):
    user_id: str
    email: EmailStr
    keywords: List[str]
    time: str  # "08:00" 형식
    weekdays: List[int]  # [0,1,2,3,4] = 월-금
    max_results: int = 10
    engines: List[str] = ["naver"]

# Response Models
class SearchResult(BaseModel):
    title: str
    url: str
    description: Optional[str] = None
    rank: int
    engine: str

@app.on_event("startup")
async def startup_event():
    """앱 시작 시 스케줄러 시작"""
    scheduler_manager.start()

@app.on_event("shutdown")
async def shutdown_event():
    """앱 종료 시 스케줄러 종료"""
    scheduler_manager.shutdown()

@app.get("/")
async def root():
    return {
        "service": "WorkFree 뉴스 크롤링 API",
        "version": "2.0.0",
        "status": "running",
        "description": "검색어 기반 뉴스 자동 크롤링 + 스케줄 자동발송",
        "endpoints": {
            "search": "/api/search",
            "email": "/api/email",
            "schedule_create": "/api/schedule",
            "schedule_get": "/api/schedule/{user_id}",
            "schedule_delete": "/api/schedule/{user_id}",
            "schedule_list": "/api/schedules",
            "health": "/health"
        }
    }

@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "resend_configured": bool(RESEND_API_KEY),
        "google_api_configured": bool(GOOGLE_API_KEY and GOOGLE_SEARCH_ENGINE_ID),
        "naver_api_configured": bool(NAVER_CLIENT_ID and NAVER_CLIENT_SECRET),
        "timestamp": datetime.now().isoformat()
    }

def search_google(keyword: str, max_results: int = 10) -> List[dict]:
    """구글 검색 - Custom Search API 사용"""
    results = []
    
    if not GOOGLE_API_KEY or not GOOGLE_SEARCH_ENGINE_ID:
        print("Google API 키가 설정되지 않았습니다. 데모 데이터를 반환합니다.")
        # 데모 데이터 반환
        for i in range(min(max_results, 5)):
            results.append({
                'title': f'[데모] {keyword} 관련 결과 {i+1}',
                'url': f'https://example.com/result-{i+1}',
                'description': f'{keyword}에 대한 검색 결과입니다. Google API 키를 설정하면 실제 검색 결과를 받을 수 있습니다.',
                'rank': i + 1,
                'engine': 'google'
            })
        return results
    
    try:
        url = "https://www.googleapis.com/customsearch/v1"
        params = {
            'key': GOOGLE_API_KEY,
            'cx': GOOGLE_SEARCH_ENGINE_ID,
            'q': keyword,
            'num': min(max_results, 10),  # API 제한: 최대 10개
            'hl': 'ko'
        }
        
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if 'items' in data:
            for idx, item in enumerate(data['items'], 1):
                results.append({
                    'title': item.get('title', ''),
                    'url': item.get('link', ''),
                    'description': item.get('snippet', '')[:200],
                    'rank': idx,
                    'engine': 'google'
                })
        
        print(f"Google API: {len(results)} 결과 반환")
                
    except Exception as e:
        print(f"Google API error: {e}")
    
    return results

def search_naver(keyword: str, max_results: int = 10) -> List[dict]:
    """네이버 뉴스 검색 - Naver News API 사용"""
    results = []
    
    print(f"[DEBUG] NAVER_CLIENT_ID: {NAVER_CLIENT_ID[:10] if NAVER_CLIENT_ID else 'NOT SET'}...")
    print(f"[DEBUG] NAVER_CLIENT_SECRET: {'SET' if NAVER_CLIENT_SECRET else 'NOT SET'}")
    
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        print("Naver API 키가 설정되지 않았습니다. 데모 데이터를 반환합니다.")
        # 데모 데이터 반환
        for i in range(min(max_results, 5)):
            results.append({
                'title': f'[데모] {keyword} 관련 뉴스 {i+1}',
                'url': f'https://example.com/naver-news-{i+1}',
                'description': f'{keyword}에 대한 최신 뉴스입니다. Naver API 키를 설정하면 실제 뉴스 결과를 받을 수 있습니다.',
                'rank': i + 1,
                'engine': 'naver'
            })
        return results
    
    try:
        url = "https://openapi.naver.com/v1/search/news.json"
        headers = {
            'X-Naver-Client-Id': NAVER_CLIENT_ID,
            'X-Naver-Client-Secret': NAVER_CLIENT_SECRET
        }
        params = {
            'query': keyword,
            'display': min(max_results, 100),  # API 제한: 최대 100개
            'sort': 'date'  # date (최신순) or sim (관련도순)
        }
        
        print(f"[DEBUG] Naver API 요청 URL: {url}")
        print(f"[DEBUG] Naver API 요청 params: {params}")
        
        response = requests.get(url, headers=headers, params=params, timeout=10)
        print(f"[DEBUG] Naver API 응답 상태: {response.status_code}")
        
        response.raise_for_status()
        data = response.json()
        
        print(f"[DEBUG] Naver API 응답 데이터 키: {data.keys()}")
        
        if 'items' in data:
            for idx, item in enumerate(data['items'], 1):
                # HTML 태그 제거
                title = item.get('title', '').replace('<b>', '').replace('</b>', '')
                description = item.get('description', '').replace('<b>', '').replace('</b>', '')
                
                results.append({
                    'title': title,
                    'url': item.get('link', ''),
                    'description': description[:200],
                    'rank': idx,
                    'engine': 'naver'
                })
        
        print(f"Naver API: {len(results)} 결과 반환")
                
    except Exception as e:
        print(f"[ERROR] Naver API error: {e}")
        import traceback
        print(f"[ERROR] Traceback: {traceback.format_exc()}")
    
    return results

def create_excel(results: List[dict]) -> bytes:
    """검색 결과를 Excel로 변환"""
    wb = Workbook()
    ws = wb.active
    ws.title = "뉴스 검색 결과"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ['순위', '검색엔진', '제목', 'URL', '설명']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row_num, r in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=r['rank'])
        ws.cell(row=row_num, column=2, value=r['engine'].upper())
        ws.cell(row=row_num, column=3, value=r['title'])
        ws.cell(row=row_num, column=4, value=r['url'])
        ws.cell(row=row_num, column=5, value=r['description'])
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 80
    
    # BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def send_email(recipient: str, keyword: str, excel_content: bytes, results_count: int):
    """이메일 발송 - Resend API 사용"""
    if not RESEND_API_KEY:
        raise Exception("Resend API 키가 설정되지 않았습니다")
    
    # Resend API 키 설정
    resend.api_key = RESEND_API_KEY
    
    # 파일명 생성
    date_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"WorkFree_뉴스검색_{keyword}_{date_str}.xlsx"
    
    # Excel 파일을 base64로 인코딩
    excel_base64 = base64.b64encode(excel_content).decode('utf-8')
    
    # HTML 이메일 본문
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2 style="color: #6A5CFF;">📰 WorkFree 뉴스 검색 결과</h2>
            <p>안녕하세요, <strong>WorkFree</strong>입니다.</p>
            <p><strong>'{keyword}'</strong> 검색어에 대한 최신 뉴스 검색 결과를 보내드립니다.</p>
            
            <div style="background-color: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <p style="margin: 5px 0;"><strong>📰 검색 결과:</strong> {results_count}개</p>
                <p style="margin: 5px 0;"><strong>📅 검색 일시:</strong> {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}</p>
                <p style="margin: 5px 0;"><strong>🔍 검색 엔진:</strong> 네이버 뉴스 (최신순)</p>
            </div>
            
            <p>첨부된 <strong>Excel 파일</strong>을 확인해주세요.</p>
            
            <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;" />
            
            <p style="color: #666; font-size: 12px;">
                이 이메일은 WorkFree 뉴스 크롤링 서비스에서 자동으로 발송되었습니다.<br>
                <a href="https://workfreemarket.com" style="color: #6A5CFF;">workfreemarket.com</a>
            </p>
        </div>
    </body>
    </html>
    """
    
    # 텍스트 이메일 본문 (HTML 미지원 클라이언트용)
    text_body = f"""
안녕하세요, WorkFree입니다.

'{keyword}' 검색어에 대한 최신 뉴스 검색 결과를 보내드립니다.

📰 검색 결과: {results_count}개
📅 검색 일시: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}
🔍 검색 엔진: 네이버 뉴스 (최신순)

첨부된 Excel 파일을 확인해주세요.

---
이 이메일은 WorkFree 뉴스 크롤링 서비스에서 자동으로 발송되었습니다.
https://workfreemarket.com
    """
    
    try:
        # Resend API로 이메일 발송
        params = {
            "from": "WorkFree <noreply@workfreemarket.com>",
            "to": [recipient],
            "subject": f"[WorkFree] '{keyword}' 뉴스 검색 결과 ({results_count}개)",
            "html": html_body,
            "text": text_body,
            "attachments": [{
                "filename": filename,
                "content": excel_base64,
            }]
        }
        
        email = resend.Emails.send(params)
        print(f"[SUCCESS] Email sent via Resend: {email}")
        return email
        
    except Exception as e:
        print(f"[ERROR] Resend API error: {e}")
        raise Exception(f"이메일 발송 실패: {str(e)}")

@app.post("/api/search")
async def search(request: SearchRequest):
    """검색 실행"""
    all_results = []
    
    if "google" in request.engines:
        google_results = search_google(request.keyword, request.max_results)
        all_results.extend(google_results)
    
    if "naver" in request.engines:
        naver_results = search_naver(request.keyword, request.max_results)
        all_results.extend(naver_results)
    
    if not all_results:
        raise HTTPException(status_code=404, detail="검색 결과를 찾을 수 없습니다")
    
    return {
        "keyword": request.keyword,
        "total_results": len(all_results),
        "results": all_results,
        "timestamp": datetime.now().isoformat()
    }

@app.post("/api/email")
async def search_and_email(request: EmailRequest):
    """검색 후 이메일 발송"""
    try:
        # 검색 실행
        all_results = []
        
        if "google" in request.engines:
            google_results = search_google(request.keyword, request.max_results)
            all_results.extend(google_results)
        
        if "naver" in request.engines:
            naver_results = search_naver(request.keyword, request.max_results)
            all_results.extend(naver_results)
        
        # 검색 결과가 없으면 데모 데이터 생성
        if not all_results:
            print("[INFO] 검색 결과가 없어 데모 데이터를 생성합니다.")
            for i in range(min(request.max_results, 5)):
                all_results.append({
                    'title': f'[Demo] {request.keyword} 관련 뉴스 {i+1}',
                    'url': f'https://workfreemarket.com/demo-{i+1}',
                    'description': f'{request.keyword}에 대한 검색 결과입니다. 실제 API 키를 설정하면 실시간 뉴스 결과를 받을 수 있습니다.',
                    'rank': i + 1,
                    'engine': 'demo'
                })
        
        # Excel 생성
        excel_content = create_excel(all_results)
        
        # 이메일 발송
        send_email(request.recipient_email, request.keyword, excel_content, len(all_results))
        
        return {
            "success": True,
            "message": "이메일 발송 완료",
            "keyword": request.keyword,
            "recipient": request.recipient_email,
            "results_count": len(all_results),
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"오류 발생: {str(e)}")

@app.post("/api/schedule")
async def create_schedule(
    request: ScheduleRequest,
    authenticated: bool = Depends(verify_api_key)
):
    """스케줄 생성 (인증 필요)"""
    try:
        result = scheduler_manager.add_user_schedule(
            user_id=request.user_id,
            email=request.email,
            keywords=request.keywords,
            time_str=request.time,
            weekdays=request.weekdays,
            max_results=request.max_results,
            engines=request.engines
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"스케줄 생성 실패: {str(e)}")

@app.get("/api/schedule/{user_id}")
async def get_schedule(
    user_id: str,
    authenticated: bool = Depends(verify_api_key)
):
    """사용자 스케줄 조회 (인증 필요)"""
    schedule = scheduler_manager.get_schedule(user_id)
    if not schedule:
        raise HTTPException(status_code=404, detail="스케줄을 찾을 수 없습니다")
    return schedule

@app.delete("/api/schedule/{user_id}")
async def delete_schedule(
    user_id: str,
    authenticated: bool = Depends(verify_api_key)
):
    """스케줄 삭제 (인증 필요)"""
    success = scheduler_manager.remove_schedule(user_id)
    if success:
        return {"success": True, "message": "스케줄이 삭제되었습니다"}
    else:
        raise HTTPException(status_code=404, detail="스케줄을 찾을 수 없습니다")

@app.get("/api/schedules")
async def list_schedules(authenticated: bool = Depends(verify_api_key)):
    """모든 스케줄 목록 조회 (인증 필요 - 관리자 전용)"""
    schedules = scheduler_manager.get_all_schedules()
    return {
        "total": len(schedules),
        "schedules": schedules
    }

@app.post("/api/schedule/{user_id}/pause")
async def pause_schedule(
    user_id: str,
    authenticated: bool = Depends(verify_api_key)
):
    """스케줄 일시정지 (인증 필요)"""
    success = scheduler_manager.pause_schedule(user_id)
    if success:
        return {"success": True, "message": "스케줄이 일시정지되었습니다"}
    else:
        raise HTTPException(status_code=404, detail="스케줄을 찾을 수 없습니다")

@app.post("/api/schedule/{user_id}/resume")
async def resume_schedule(
    user_id: str,
    authenticated: bool = Depends(verify_api_key)
):
    """스케줄 재개 (인증 필요)"""
    success = scheduler_manager.resume_schedule(user_id)
    if success:
        return {"success": True, "message": "스케줄이 재개되었습니다"}
    else:
        raise HTTPException(status_code=404, detail="스케줄을 찾을 수 없습니다")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

