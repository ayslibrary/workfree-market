import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import winreg

# 엑셀 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "인수인계 할 일 목록"

# 스타일 정의
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 카테고리별 색상
category_colors = {
    "시스템/프로그램": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "업무 프로세스": PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"),
    "고객사/거래처": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "문서/매뉴얼": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
    "기타": PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
}

# 헤더 작성
headers = ["카테고리", "할 일 항목", "담당자", "우선순위", "예상소요시간", "상세내용", "완료여부", "비고"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# 인수인계 할 일 목록 (이미지에서 확인한 내용 기반)
tasks = [
    # 시스템/프로그램 관련
    {
        "category": "시스템/프로그램",
        "task": "STARNET 등 설치파일 설치",
        "person": "정인님",
        "priority": "높음",
        "duration": "1시간",
        "detail": "STARNET 등 필요한 프로그램 설치 및 설정 확인",
        "status": "",
        "note": "설치 파일 위치 및 설치 순서 매뉴얼 준비"
    },
    {
        "category": "시스템/프로그램",
        "task": "서명 만들기",
        "person": "정인님",
        "priority": "중",
        "duration": "30분",
        "detail": "이메일 서명 파일 생성 및 설정",
        "status": "",
        "note": "서명 템플릿 파일 전달 필요"
    },
    
    # 업무 프로세스 관련
    {
        "category": "업무 프로세스",
        "task": "정인님 인수인계 일정 및 내용 설명",
        "person": "담당자",
        "priority": "높음",
        "duration": "1시간",
        "detail": "전체 인수인계 일정 및 각 항목별 내용 설명",
        "status": "",
        "note": "첫날 오전에 진행"
    },
    
    # 고객사/거래처 관련
    {
        "category": "고객사/거래처",
        "task": "침퓨즈(APRO, DCC) 인수인계",
        "person": "영화 과장님",
        "priority": "높음",
        "duration": "40분",
        "detail": "침퓨즈 관련 APRO, DCC 업무 프로세스 및 주요 사항 인수인계",
        "status": "",
        "note": "출근자: 영화, 송은, 현정"
    },
    {
        "category": "고객사/거래처",
        "task": "악세스(ACCSS-노이즈필터) 인수인계",
        "person": "승은 차장님, 현정 과장님",
        "priority": "높음",
        "duration": "20분",
        "detail": "ACCSS 노이즈필터 관련 업무 프로세스 인수인계",
        "status": "",
        "note": "정인님 OJT"
    },
    {
        "category": "고객사/거래처",
        "task": "커넥터 EMS 3사 전체적으로 인수인계 설명",
        "person": "미경 선배님",
        "priority": "높음",
        "duration": "1시간",
        "detail": "커넥터 EMS 3개사 전체적인 업무 프로세스 및 주요 사항 설명",
        "status": "",
        "note": "출근자: 미경, 현정"
    },
    {
        "category": "고객사/거래처",
        "task": "한국성전 - FCST 인계",
        "person": "미경 선배님",
        "priority": "중",
        "duration": "30분",
        "detail": "한국성전 FCST 관련 업무 인수인계",
        "status": "",
        "note": "출근자: 미경, 현정"
    },
    {
        "category": "고객사/거래처",
        "task": "샘플 인계 - 유상 및 발주 부분",
        "person": "담당자",
        "priority": "중",
        "duration": "1시간",
        "detail": "샘플 관련 유상 및 발주 업무 프로세스 인수인계",
        "status": "",
        "note": "출근자: 송은"
    },
    {
        "category": "고객사/거래처",
        "task": "샘플 - 발주 이레귤러 바뀐 부분 인계",
        "person": "미경 선배님",
        "priority": "중",
        "duration": "30분",
        "detail": "샘플 발주 중 이레귤러하게 변경된 부분에 대한 인수인계",
        "status": "",
        "note": "출근자: 송은, 현정, 미경"
    },
    
    # 문서/매뉴얼 관련
    {
        "category": "문서/매뉴얼",
        "task": "매뉴얼 파일 정리 및 전달",
        "person": "담당자",
        "priority": "높음",
        "duration": "2시간",
        "detail": "정리한 매뉴얼 파일을 정인님에게 전달하여 업무 수행하게 하기",
        "status": "",
        "note": "각 업무별 매뉴얼 파일 체계적으로 정리 필요"
    },
    {
        "category": "문서/매뉴얼",
        "task": "할 일 시트 링크 생성",
        "person": "담당자",
        "priority": "낮음",
        "duration": "30분",
        "detail": "옆에 할 일에 대한 시트로 넘어가는 링크 만들기",
        "status": "",
        "note": "엑셀 하이퍼링크 기능 활용"
    },
    
    # 기타
    {
        "category": "기타",
        "task": "출근자 일정 확인 및 조율",
        "person": "담당자",
        "priority": "중",
        "duration": "1시간",
        "detail": "각 인수인계 항목별 필요한 출근자 일정 확인 및 조율",
        "status": "",
        "note": "재택근무자와 출근자 일정 매칭 필요"
    },
    {
        "category": "기타",
        "task": "인수인계 완료 체크리스트 작성",
        "person": "담당자",
        "priority": "중",
        "duration": "1시간",
        "detail": "각 항목별 인수인계 완료 여부 확인용 체크리스트 작성",
        "status": "",
        "note": "실제 업무 수행 가능 여부 확인"
    },
    {
        "category": "기타",
        "task": "Q&A 시간 확보",
        "person": "전체",
        "priority": "중",
        "duration": "매일 30분",
        "detail": "인수인계 후 궁금한 사항 질문 및 답변 시간",
        "status": "",
        "note": "매일 하루 종료 전 정리 시간 활용"
    }
]

# 할 일 목록 작성
row = 2
for task in tasks:
    # 카테고리
    cat_cell = ws.cell(row=row, column=1, value=task["category"])
    cat_cell.fill = category_colors.get(task["category"], category_colors["기타"])
    cat_cell.alignment = Alignment(horizontal='center', vertical='center')
    cat_cell.border = border_style
    cat_cell.font = Font(size=10)
    
    # 할 일 항목
    task_cell = ws.cell(row=row, column=2, value=task["task"])
    task_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    task_cell.border = border_style
    task_cell.font = Font(size=10, bold=True)
    
    # 담당자
    person_cell = ws.cell(row=row, column=3, value=task["person"])
    person_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    person_cell.border = border_style
    person_cell.font = Font(size=10)
    
    # 우선순위
    priority_cell = ws.cell(row=row, column=4, value=task["priority"])
    priority_cell.alignment = Alignment(horizontal='center', vertical='center')
    priority_cell.border = border_style
    priority_cell.font = Font(size=10)
    # 우선순위에 따른 색상
    if task["priority"] == "높음":
        priority_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif task["priority"] == "중":
        priority_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        priority_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # 예상소요시간
    duration_cell = ws.cell(row=row, column=5, value=task["duration"])
    duration_cell.alignment = Alignment(horizontal='center', vertical='center')
    duration_cell.border = border_style
    duration_cell.font = Font(size=10)
    
    # 상세내용
    detail_cell = ws.cell(row=row, column=6, value=task["detail"])
    detail_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    detail_cell.border = border_style
    detail_cell.font = Font(size=10)
    
    # 완료여부
    status_cell = ws.cell(row=row, column=7, value=task["status"])
    status_cell.alignment = Alignment(horizontal='center', vertical='center')
    status_cell.border = border_style
    status_cell.font = Font(size=10)
    
    # 비고
    note_cell = ws.cell(row=row, column=8, value=task["note"])
    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    note_cell.border = border_style
    note_cell.font = Font(size=9, italic=True)
    
    row += 1

# 열 너비 조정
column_widths = {
    'A': 18,  # 카테고리
    'B': 35,  # 할 일 항목
    'C': 20,  # 담당자
    'D': 12,  # 우선순위
    'E': 15,  # 예상소요시간
    'F': 40,  # 상세내용
    'G': 12,  # 완료여부
    'H': 30   # 비고
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 행 높이 조정
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 40

# 첫 번째 행 높이 조정
ws.row_dimensions[1].height = 40

# 완료여부 컬럼에 데이터 유효성 검사 추가 (드롭다운)
from openpyxl.worksheet.datavalidation import DataValidation

status_validation = DataValidation(type="list", formula1='"미완료,진행중,완료,보류"', allow_blank=True)
status_validation.error = "목록에서 선택해주세요"
status_validation.errorTitle = "잘못된 입력"
status_validation.prompt = "완료 상태를 선택하세요"
status_validation.promptTitle = "완료여부 선택"

ws.add_data_validation(status_validation)
status_validation.add(f'G2:G{ws.max_row}')

# 필터 추가
ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

# 요약 시트 생성
summary_ws = wb.create_sheet("요약")
summary_ws.append(["카테고리", "항목 수", "완료", "진행중", "미완료"])
summary_ws.append(["시스템/프로그램", 2, 0, 0, 2])
summary_ws.append(["업무 프로세스", 1, 0, 0, 1])
summary_ws.append(["고객사/거래처", 5, 0, 0, 5])
summary_ws.append(["문서/매뉴얼", 2, 0, 0, 2])
summary_ws.append(["기타", 3, 0, 0, 3])
summary_ws.append(["합계", 13, 0, 0, 13])

# 요약 시트 스타일 적용
for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row):
    for cell in row:
        cell.border = border_style
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if cell.row == 1:
            cell.fill = header_fill
            cell.font = header_font

# 파일 저장 (바탕화면에 저장)
try:
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
    desktop_path = winreg.QueryValueEx(key, "Desktop")[0]
    winreg.CloseKey(key)
except:
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    if not os.path.exists(desktop_path):
        desktop_path = os.path.join(os.path.expanduser('~'), 'Documents')

filename = os.path.join(desktop_path, "인수인계_할일목록.xlsx")
wb.save(filename)
print(f"인수인계 할 일 목록이 바탕화면에 생성되었습니다: {filename}")
print(f"\n총 {len(tasks)}개의 할 일 항목이 정리되었습니다.")
print("\n카테고리별 분류:")
for category in category_colors.keys():
    count = len([t for t in tasks if t["category"] == category])
    print(f"  - {category}: {count}개")

