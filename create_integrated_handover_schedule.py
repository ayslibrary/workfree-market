import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from openpyxl.worksheet.datavalidation import DataValidation
import os
import winreg

# 날짜 범위 설정
start_date = datetime(2025, 11, 17)
end_date = datetime(2025, 11, 28)

# 엑셀 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "인수인계 스케줄"

# 스타일 정의
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
date_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
lunch_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# 개선된 헤더 구조
headers = [
    "날짜", 
    "요일", 
    "시간", 
    "주제/내용", 
    "담당자", 
    "소요시간", 
    "상세내용", 
    "출근자", 
    "재택근무자", 
    "비고", 
    "완료여부"
]

# 헤더 작성
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# 날짜별로 행 생성
row = 2
current_date = start_date
weekdays = ['월', '화', '수', '목', '금', '토', '일']

while current_date <= end_date:
    weekday = weekdays[current_date.weekday()]
    
    # 시간대별 행 생성 (9시~18시, 1시간 단위)
    time_slots = []
    for hour in range(9, 18):
        time_str = f"{hour:02d}:00 - {hour+1:02d}:00"
        time_slots.append(time_str)
    
    # 점심시간 처리
    lunch_time_idx = 3  # 12:00-13:00
    
    for time_idx, time_slot in enumerate(time_slots):
        current_row = row + time_idx
        
        # 날짜와 요일 (첫 번째 시간대에만)
        if time_idx == 0:
            date_cell = ws.cell(row=current_row, column=1, value=current_date.strftime("%Y-%m-%d"))
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            date_cell.border = border_style
            date_cell.font = Font(bold=True, size=10)
            date_cell.fill = date_fill
            
            weekday_cell = ws.cell(row=current_row, column=2, value=weekday)
            weekday_cell.alignment = Alignment(horizontal='center', vertical='center')
            weekday_cell.border = border_style
            weekday_cell.fill = date_fill
        else:
            # 나머지 시간대는 빈 셀 (병합을 위해)
            ws.cell(row=current_row, column=1, value="").border = border_style
            ws.cell(row=current_row, column=2, value="").border = border_style
        
        # 점심시간 스타일 적용
        is_lunch = (time_idx == lunch_time_idx)
        
        # 시간
        time_cell = ws.cell(row=current_row, column=3, value=time_slot)
        time_cell.alignment = Alignment(horizontal='center', vertical='center')
        time_cell.border = border_style
        if is_lunch:
            time_cell.fill = lunch_fill
            time_cell.value = "12:00 - 13:00 (점심시간)"
        
        # 주제/내용
        content_cell = ws.cell(row=current_row, column=4, value="")
        content_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        content_cell.border = border_style
        if is_lunch:
            content_cell.fill = lunch_fill
        
        # 담당자
        person_cell = ws.cell(row=current_row, column=5, value="")
        person_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        person_cell.border = border_style
        if is_lunch:
            person_cell.fill = lunch_fill
        
        # 소요시간
        duration_cell = ws.cell(row=current_row, column=6, value="")
        duration_cell.alignment = Alignment(horizontal='center', vertical='center')
        duration_cell.border = border_style
        if is_lunch:
            duration_cell.fill = lunch_fill
        
        # 상세내용
        detail_cell = ws.cell(row=current_row, column=7, value="")
        detail_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        detail_cell.border = border_style
        if is_lunch:
            detail_cell.fill = lunch_fill
        
        # 출근자
        attendee_cell = ws.cell(row=current_row, column=8, value="")
        attendee_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        attendee_cell.border = border_style
        if is_lunch:
            attendee_cell.fill = lunch_fill
        
        # 재택근무자
        wfh_cell = ws.cell(row=current_row, column=9, value="")
        wfh_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wfh_cell.border = border_style
        if is_lunch:
            wfh_cell.fill = lunch_fill
        
        # 비고
        note_cell = ws.cell(row=current_row, column=10, value="")
        note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        note_cell.border = border_style
        if is_lunch:
            note_cell.fill = lunch_fill
        
        # 완료여부 (드롭다운을 위한 빈 셀)
        status_cell = ws.cell(row=current_row, column=11, value="")
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.border = border_style
        if is_lunch:
            status_cell.fill = lunch_fill
    
    # 날짜와 요일 셀 병합
    if len(time_slots) > 1:
        ws.merge_cells(f'A{row}:A{row + len(time_slots) - 1}')
        ws.merge_cells(f'B{row}:B{row + len(time_slots) - 1}')
    
    row += len(time_slots)
    current_date += timedelta(days=1)

# 열 너비 조정
column_widths = {
    'A': 12,  # 날짜
    'B': 8,   # 요일
    'C': 18,  # 시간
    'D': 35,  # 주제/내용
    'E': 20,  # 담당자
    'F': 12,  # 소요시간
    'G': 40,  # 상세내용
    'H': 18,  # 출근자
    'I': 18,  # 재택근무자
    'J': 25,  # 비고
    'K': 12   # 완료여부
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 행 높이 조정
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 30

# 첫 번째 행 높이 조정
ws.row_dimensions[1].height = 35

# 완료여부 컬럼에 데이터 유효성 검사 추가 (드롭다운)
status_validation = DataValidation(type="list", formula1='"미완료,진행중,완료,보류"', allow_blank=True)
status_validation.error = "목록에서 선택해주세요"
status_validation.errorTitle = "잘못된 입력"
status_validation.prompt = "완료 상태를 선택하세요"
status_validation.promptTitle = "완료여부 선택"

ws.add_data_validation(status_validation)
status_validation.add(f'K2:K{ws.max_row}')

# 필터 추가
ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

# 보기 옵션 설정
ws.sheet_view.showGridLines = True
ws.sheet_view.zoomScale = 90

# ========== 할 일 목록 시트 추가 ==========
todo_ws = wb.create_sheet("인수인계 할 일 목록")

# 스타일 정의 (할 일 목록용)
todo_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
todo_header_font = Font(bold=True, color="FFFFFF", size=12)

# 카테고리별 색상
category_colors = {
    "시스템/프로그램": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "업무 프로세스": PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"),
    "고객사/거래처": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "문서/매뉴얼": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
    "기타": PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
}

# 헤더 작성
todo_headers = ["카테고리", "할 일 항목", "담당자", "우선순위", "예상소요시간", "상세내용", "완료여부", "비고"]
for col_idx, header in enumerate(todo_headers, start=1):
    cell = todo_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = todo_header_fill
    cell.font = todo_header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# 인수인계 할 일 목록
tasks = [
    # 시스템/프로그램 관련
    {
        "category": "시스템/프로그램",
        "task": "STARNET 등 설치파일 설치",
        "person": "정인님",
        "priority": "높음",
        "duration": "1시간",
        "detail": "STARNET 등 필요한 프로그램 설치 및 설정 확인",
        "status": "",
        "note": "설치 파일 위치 및 설치 순서 매뉴얼 준비"
    },
    {
        "category": "시스템/프로그램",
        "task": "서명 만들기",
        "person": "정인님",
        "priority": "중",
        "duration": "30분",
        "detail": "이메일 서명 파일 생성 및 설정",
        "status": "",
        "note": "서명 템플릿 파일 전달 필요"
    },
    
    # 업무 프로세스 관련
    {
        "category": "업무 프로세스",
        "task": "정인님 인수인계 일정 및 내용 설명",
        "person": "담당자",
        "priority": "높음",
        "duration": "1시간",
        "detail": "전체 인수인계 일정 및 각 항목별 내용 설명",
        "status": "",
        "note": "첫날 오전에 진행"
    },
    
    # 고객사/거래처 관련
    {
        "category": "고객사/거래처",
        "task": "침퓨즈(APRO, DCC) 인수인계",
        "person": "영화 과장님",
        "priority": "높음",
        "duration": "40분",
        "detail": "침퓨즈 관련 APRO, DCC 업무 프로세스 및 주요 사항 인수인계",
        "status": "",
        "note": "출근자: 영화, 송은, 현정"
    },
    {
        "category": "고객사/거래처",
        "task": "악세스(ACCSS-노이즈필터) 인수인계",
        "person": "승은 차장님, 현정 과장님",
        "priority": "높음",
        "duration": "20분",
        "detail": "ACCSS 노이즈필터 관련 업무 프로세스 인수인계",
        "status": "",
        "note": "정인님 OJT"
    },
    {
        "category": "고객사/거래처",
        "task": "커넥터 EMS 3사 전체적으로 인수인계 설명",
        "person": "미경 선배님",
        "priority": "높음",
        "duration": "1시간",
        "detail": "커넥터 EMS 3개사 전체적인 업무 프로세스 및 주요 사항 설명",
        "status": "",
        "note": "출근자: 미경, 현정"
    },
    {
        "category": "고객사/거래처",
        "task": "한국성전 - FCST 인계",
        "person": "미경 선배님",
        "priority": "중",
        "duration": "30분",
        "detail": "한국성전 FCST 관련 업무 인수인계",
        "status": "",
        "note": "출근자: 미경, 현정"
    },
    {
        "category": "고객사/거래처",
        "task": "샘플 인계 - 유상 및 발주 부분",
        "person": "담당자",
        "priority": "중",
        "duration": "1시간",
        "detail": "샘플 관련 유상 및 발주 업무 프로세스 인수인계",
        "status": "",
        "note": "출근자: 송은"
    },
    {
        "category": "고객사/거래처",
        "task": "샘플 - 발주 이레귤러 바뀐 부분 인계",
        "person": "미경 선배님",
        "priority": "중",
        "duration": "30분",
        "detail": "샘플 발주 중 이레귤러하게 변경된 부분에 대한 인수인계",
        "status": "",
        "note": "출근자: 송은, 현정, 미경"
    },
    
    # 문서/매뉴얼 관련
    {
        "category": "문서/매뉴얼",
        "task": "매뉴얼 파일 정리 및 전달",
        "person": "담당자",
        "priority": "높음",
        "duration": "2시간",
        "detail": "정리한 매뉴얼 파일을 정인님에게 전달하여 업무 수행하게 하기",
        "status": "",
        "note": "각 업무별 매뉴얼 파일 체계적으로 정리 필요"
    },
    {
        "category": "문서/매뉴얼",
        "task": "할 일 시트 링크 생성",
        "person": "담당자",
        "priority": "낮음",
        "duration": "30분",
        "detail": "옆에 할 일에 대한 시트로 넘어가는 링크 만들기",
        "status": "",
        "note": "엑셀 하이퍼링크 기능 활용"
    },
    
    # 기타
    {
        "category": "기타",
        "task": "출근자 일정 확인 및 조율",
        "person": "담당자",
        "priority": "중",
        "duration": "1시간",
        "detail": "각 인수인계 항목별 필요한 출근자 일정 확인 및 조율",
        "status": "",
        "note": "재택근무자와 출근자 일정 매칭 필요"
    },
    {
        "category": "기타",
        "task": "인수인계 완료 체크리스트 작성",
        "person": "담당자",
        "priority": "중",
        "duration": "1시간",
        "detail": "각 항목별 인수인계 완료 여부 확인용 체크리스트 작성",
        "status": "",
        "note": "실제 업무 수행 가능 여부 확인"
    },
    {
        "category": "기타",
        "task": "Q&A 시간 확보",
        "person": "전체",
        "priority": "중",
        "duration": "매일 30분",
        "detail": "인수인계 후 궁금한 사항 질문 및 답변 시간",
        "status": "",
        "note": "매일 하루 종료 전 정리 시간 활용"
    }
]

# 할 일 목록 작성
todo_row = 2
for task in tasks:
    # 카테고리
    cat_cell = todo_ws.cell(row=todo_row, column=1, value=task["category"])
    cat_cell.fill = category_colors.get(task["category"], category_colors["기타"])
    cat_cell.alignment = Alignment(horizontal='center', vertical='center')
    cat_cell.border = border_style
    cat_cell.font = Font(size=10)
    
    # 할 일 항목
    task_cell = todo_ws.cell(row=todo_row, column=2, value=task["task"])
    task_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    task_cell.border = border_style
    task_cell.font = Font(size=10, bold=True)
    
    # 담당자
    person_cell = todo_ws.cell(row=todo_row, column=3, value=task["person"])
    person_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    person_cell.border = border_style
    person_cell.font = Font(size=10)
    
    # 우선순위
    priority_cell = todo_ws.cell(row=todo_row, column=4, value=task["priority"])
    priority_cell.alignment = Alignment(horizontal='center', vertical='center')
    priority_cell.border = border_style
    priority_cell.font = Font(size=10)
    # 우선순위에 따른 색상
    if task["priority"] == "높음":
        priority_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif task["priority"] == "중":
        priority_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        priority_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # 예상소요시간
    duration_cell = todo_ws.cell(row=todo_row, column=5, value=task["duration"])
    duration_cell.alignment = Alignment(horizontal='center', vertical='center')
    duration_cell.border = border_style
    duration_cell.font = Font(size=10)
    
    # 상세내용
    detail_cell = todo_ws.cell(row=todo_row, column=6, value=task["detail"])
    detail_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    detail_cell.border = border_style
    detail_cell.font = Font(size=10)
    
    # 완료여부
    status_cell = todo_ws.cell(row=todo_row, column=7, value=task["status"])
    status_cell.alignment = Alignment(horizontal='center', vertical='center')
    status_cell.border = border_style
    status_cell.font = Font(size=10)
    
    # 비고
    note_cell = todo_ws.cell(row=todo_row, column=8, value=task["note"])
    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    note_cell.border = border_style
    note_cell.font = Font(size=9, italic=True)
    
    todo_row += 1

# 할 일 목록 열 너비 조정
todo_column_widths = {
    'A': 18,  # 카테고리
    'B': 35,  # 할 일 항목
    'C': 20,  # 담당자
    'D': 12,  # 우선순위
    'E': 15,  # 예상소요시간
    'F': 40,  # 상세내용
    'G': 12,  # 완료여부
    'H': 30   # 비고
}

for col, width in todo_column_widths.items():
    todo_ws.column_dimensions[col].width = width

# 할 일 목록 행 높이 조정
for row in range(2, todo_ws.max_row + 1):
    todo_ws.row_dimensions[row].height = 40

# 할 일 목록 첫 번째 행 높이 조정
todo_ws.row_dimensions[1].height = 40

# 할 일 목록 완료여부 컬럼에 데이터 유효성 검사 추가 (드롭다운)
todo_status_validation = DataValidation(type="list", formula1='"미완료,진행중,완료,보류"', allow_blank=True)
todo_status_validation.error = "목록에서 선택해주세요"
todo_status_validation.errorTitle = "잘못된 입력"
todo_status_validation.prompt = "완료 상태를 선택하세요"
todo_status_validation.promptTitle = "완료여부 선택"

todo_ws.add_data_validation(todo_status_validation)
todo_status_validation.add(f'G2:G{todo_ws.max_row}')

# 할 일 목록 필터 추가
todo_ws.auto_filter.ref = f'A1:{get_column_letter(todo_ws.max_column)}{todo_ws.max_row}'

# 스케줄 시트와 할 일 목록 시트 간 하이퍼링크 추가
# 스케줄 시트의 첫 번째 셀에 할 일 목록으로 가는 링크 추가
link_cell = ws.cell(row=1, column=1)
link_cell.hyperlink = f"#'인수인계 할 일 목록'!A1"
link_cell.value = "📋 할 일 목록 보기"
link_cell.font = Font(color="0000FF", underline="single")
link_cell.alignment = Alignment(horizontal='left', vertical='center')

# 할 일 목록 시트의 첫 번째 셀에 스케줄로 가는 링크 추가
todo_link_cell = todo_ws.cell(row=1, column=1)
todo_link_cell.hyperlink = f"#'인수인계 스케줄'!A1"
todo_link_cell.value = "📅 스케줄 보기"
todo_link_cell.font = Font(color="0000FF", underline="single")
todo_link_cell.alignment = Alignment(horizontal='left', vertical='center')

# 파일 저장 (바탕화면에 저장)
try:
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
    desktop_path = winreg.QueryValueEx(key, "Desktop")[0]
    winreg.CloseKey(key)
except:
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    if not os.path.exists(desktop_path):
        desktop_path = os.path.join(os.path.expanduser('~'), 'Documents')

filename = os.path.join(desktop_path, "신입_인수인계_통합.xlsx")
wb.save(filename)
print(f"통합 엑셀 파일이 바탕화면에 생성되었습니다: {filename}")
print(f"\n포함된 시트:")
print(f"  1. 인수인계 스케줄 - 날짜별 시간대별 스케줄표")
print(f"  2. 인수인계 할 일 목록 - 카테고리별 할 일 목록 ({len(tasks)}개 항목)")
print(f"\n각 시트의 첫 번째 셀에서 다른 시트로 이동할 수 있는 링크가 추가되었습니다.")

