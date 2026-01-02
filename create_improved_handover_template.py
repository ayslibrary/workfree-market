import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import winreg

# 날짜 범위 설정
start_date = datetime(2025, 11, 17)
end_date = datetime(2025, 11, 28)

# 엑셀 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "인수인계 스케줄"

# 스타일 정의
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
date_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
lunch_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# 개선된 헤더 구조
headers = [
    "날짜", 
    "요일", 
    "시간", 
    "주제/내용", 
    "담당자", 
    "소요시간", 
    "상세내용", 
    "출근자", 
    "재택근무자", 
    "비고", 
    "완료여부"
]

# 헤더 작성
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# 날짜별로 행 생성
row = 2
current_date = start_date
weekdays = ['월', '화', '수', '목', '금', '토', '일']

while current_date <= end_date:
    weekday = weekdays[current_date.weekday()]
    
    # 시간대별 행 생성 (9시~18시, 1시간 단위)
    time_slots = []
    for hour in range(9, 18):
        time_str = f"{hour:02d}:00 - {hour+1:02d}:00"
        time_slots.append(time_str)
    
    # 점심시간 처리
    lunch_time_idx = 3  # 12:00-13:00
    
    for time_idx, time_slot in enumerate(time_slots):
        current_row = row + time_idx
        
        # 날짜와 요일 (첫 번째 시간대에만)
        if time_idx == 0:
            date_cell = ws.cell(row=current_row, column=1, value=current_date.strftime("%Y-%m-%d"))
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            date_cell.border = border_style
            date_cell.font = Font(bold=True, size=10)
            date_cell.fill = date_fill
            
            weekday_cell = ws.cell(row=current_row, column=2, value=weekday)
            weekday_cell.alignment = Alignment(horizontal='center', vertical='center')
            weekday_cell.border = border_style
            weekday_cell.fill = date_fill
        else:
            # 나머지 시간대는 빈 셀 (병합을 위해)
            ws.cell(row=current_row, column=1, value="").border = border_style
            ws.cell(row=current_row, column=2, value="").border = border_style
        
        # 점심시간 스타일 적용
        is_lunch = (time_idx == lunch_time_idx)
        
        # 시간
        time_cell = ws.cell(row=current_row, column=3, value=time_slot)
        time_cell.alignment = Alignment(horizontal='center', vertical='center')
        time_cell.border = border_style
        if is_lunch:
            time_cell.fill = lunch_fill
            time_cell.value = "12:00 - 13:00 (점심시간)"
        
        # 주제/내용
        content_cell = ws.cell(row=current_row, column=4, value="")
        content_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        content_cell.border = border_style
        if is_lunch:
            content_cell.fill = lunch_fill
        
        # 담당자
        person_cell = ws.cell(row=current_row, column=5, value="")
        person_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        person_cell.border = border_style
        if is_lunch:
            person_cell.fill = lunch_fill
        
        # 소요시간
        duration_cell = ws.cell(row=current_row, column=6, value="")
        duration_cell.alignment = Alignment(horizontal='center', vertical='center')
        duration_cell.border = border_style
        if is_lunch:
            duration_cell.fill = lunch_fill
        
        # 상세내용
        detail_cell = ws.cell(row=current_row, column=7, value="")
        detail_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        detail_cell.border = border_style
        if is_lunch:
            detail_cell.fill = lunch_fill
        
        # 출근자
        attendee_cell = ws.cell(row=current_row, column=8, value="")
        attendee_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        attendee_cell.border = border_style
        if is_lunch:
            attendee_cell.fill = lunch_fill
        
        # 재택근무자
        wfh_cell = ws.cell(row=current_row, column=9, value="")
        wfh_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wfh_cell.border = border_style
        if is_lunch:
            wfh_cell.fill = lunch_fill
        
        # 비고
        note_cell = ws.cell(row=current_row, column=10, value="")
        note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        note_cell.border = border_style
        if is_lunch:
            note_cell.fill = lunch_fill
        
        # 완료여부 (드롭다운을 위한 빈 셀)
        status_cell = ws.cell(row=current_row, column=11, value="")
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.border = border_style
        if is_lunch:
            status_cell.fill = lunch_fill
    
    # 날짜와 요일 셀 병합
    if len(time_slots) > 1:
        ws.merge_cells(f'A{row}:A{row + len(time_slots) - 1}')
        ws.merge_cells(f'B{row}:B{row + len(time_slots) - 1}')
    
    row += len(time_slots)
    current_date += timedelta(days=1)

# 열 너비 조정
column_widths = {
    'A': 12,  # 날짜
    'B': 8,   # 요일
    'C': 18,  # 시간
    'D': 35,  # 주제/내용
    'E': 20,  # 담당자
    'F': 12,  # 소요시간
    'G': 40,  # 상세내용
    'H': 18,  # 출근자
    'I': 18,  # 재택근무자
    'J': 25,  # 비고
    'K': 12   # 완료여부
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 행 높이 조정
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 30

# 첫 번째 행 높이 조정
ws.row_dimensions[1].height = 35

# 완료여부 컬럼에 데이터 유효성 검사 추가 (드롭다운)
from openpyxl.worksheet.datavalidation import DataValidation

status_validation = DataValidation(type="list", formula1='"미완료,진행중,완료,보류"', allow_blank=True)
status_validation.error = "목록에서 선택해주세요"
status_validation.errorTitle = "잘못된 입력"
status_validation.prompt = "완료 상태를 선택하세요"
status_validation.promptTitle = "완료여부 선택"

ws.add_data_validation(status_validation)
status_validation.add(f'K2:K{ws.max_row}')

# 필터 추가
ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

# 보기 옵션 설정
ws.sheet_view.showGridLines = True
ws.sheet_view.zoomScale = 90

# 파일 저장 (바탕화면에 저장)
try:
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
    desktop_path = winreg.QueryValueEx(key, "Desktop")[0]
    winreg.CloseKey(key)
except:
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    if not os.path.exists(desktop_path):
        desktop_path = os.path.join(os.path.expanduser('~'), 'Documents')

filename = os.path.join(desktop_path, "신입_인수인계_스케줄표_개선버전.xlsx")
wb.save(filename)
print(f"개선된 엑셀 파일이 바탕화면에 생성되었습니다: {filename}")
print("\n개선 사항:")
print("1. 출근자와 재택근무자 컬럼 분리")
print("2. 완료여부 드롭다운 추가 (미완료/진행중/완료/보류)")
print("3. 필터 기능 활성화")
print("4. 점심시간 시각적 구분")
print("5. 더 명확한 컬럼 구조")

